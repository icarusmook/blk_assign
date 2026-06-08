"""
blk_assign_agent.py — 블록 작업장 배정 시스템 + HTML 리포트 자동 생성

[글로벌 규칙]
  · load_mh를 m_stdt~m_fndt 영업일 균등 분배 → 월별/주별 부하 산출 (전 규칙 공통)
  · 전체 조업도(부하/능력) 균등화 목표 / 인접 주 여유도 반영
  · 결과: assigned_area 컬럼 추가 → blk_assign_result.xlsx

[순차 배정 규칙]
  R0.  이관물량(H) 처리가능여부 리포팅
  R1.  작업장별 월별 배정 목표물량 산출
  R2.  H/Y9 블록 → AREA15 (목표물량 무관)
  R3.  T-중/대   → m_area
  R4.  T-소/동일 → prt_area
  R5.  H-소 AREA3~5 처리가능여부 리포팅
  R6.  H-소 → AREA3~5 순환(round-robin)
  R7.  H-중-jig_F → AREA2 (Part1 특정blk 무조건 + Part2 목표까지)
  R8.  H-중-jig_D/L/W → AREA1→2→6 순환
  R9.  H-전문화 family → AREA7/8/9/10
  R10. H-대-jig_D/L 그룹 → AREA12
  R11. H-대/중 잔여 → AREA7~14 (PS페어 우선, 불가시 개별)
  R12. H-dir(blk[3]=9/H) → prt_area (override)
  R13. T-소 분산배정: area_seq 순, (pjt+blk+type) 그룹, assign_prior→m_stdt 정렬
"""

import sys
import json
import pandas as pd
import numpy as np
from datetime import date, timedelta, datetime
from collections import defaultdict
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

# ── 경로 ─────────────────────────────────────────────────────────────────────
BASE_DIR        = r'D:\AI\blk_assign'
BLK_MASTER_PATH = f'{BASE_DIR}/blk_master.xlsx'
AREA_CAPA_PATH  = f'{BASE_DIR}/area_capa.xlsx'
RESULT_PATH     = f'{BASE_DIR}/blk_assign_result.xlsx'
STATS_PATH      = f'{BASE_DIR}/blk_assign_stats.json'
REPORT_PATH     = f'{BASE_DIR}/blk_assign_report.html'

# ── 규칙 추적 ─────────────────────────────────────────────────────────────────
_current_rule: str = ''

def set_rule(rule: str):
    global _current_rule
    _current_rule = rule

# ── 공휴일 ───────────────────────────────────────────────────────────────────
KR_HOLIDAYS = {
    date(2026,  1,  1),
    date(2026,  2, 16), date(2026,  2, 17), date(2026,  2, 18),
    date(2026,  3,  1),
    date(2026,  5,  5), date(2026,  5, 26),
    date(2026,  6,  6),
    date(2026,  8, 15),
    date(2026,  9, 25), date(2026,  9, 26), date(2026,  9, 27), date(2026,  9, 28),
    date(2026, 10,  3), date(2026, 10,  9),
    date(2026, 12, 25),
    date(2027,  1,  1), date(2027,  1, 27), date(2027,  1, 28), date(2027,  1, 29),
    date(2027,  3,  1),
}

# ── 작업장 그룹 상수 ──────────────────────────────────────────────────────────
AREA_Y9       = 'AREA15'
AREAS_SMALL   = ['AREA3', 'AREA4', 'AREA5']
AREAS_MID_DL  = ['AREA1', 'AREA2', 'AREA6']
AREAS_REMAIN  = ['AREA7', 'AREA8', 'AREA9', 'AREA10', 'AREA12', 'AREA13', 'AREA14']
PS_AREA_PAIRS = [('AREA7', 'AREA8'), ('AREA9', 'AREA10'), ('AREA13', 'AREA14')]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 유틸리티
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def is_workday(d: date) -> bool:
    return d.weekday() < 5 and d not in KR_HOLIDAYS


def get_workdays_between(start: date, end: date) -> list:
    days, cur = [], start
    while cur <= end:
        if is_workday(cur):
            days.append(cur)
        cur += timedelta(days=1)
    return days


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 데이터 로드
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def load_blk_master(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    df['m_stdt'] = pd.to_datetime(df['m_stdt'])
    df['m_fndt'] = pd.to_datetime(df['m_fndt'])
    df['assigned_area'] = None
    df['rule_assigned']  = ''
    return df


def load_area_capa(path: str) -> dict:
    """반환: {area_name: {month_int: capacity_mh}}"""
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    capa = {}
    for row in rows[2:]:
        area = row[0]
        if not area:
            continue
        capa[area] = {m: (row[m] or 0) for m in range(1, 13)}
    return capa


def load_area_seq(path: str) -> list:
    """area_capa.xlsx col[13](assign_prior) 오름차순으로 정렬된 area 목록 반환"""
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    areas = []
    for row in rows[2:]:
        if row[0]:
            seq = row[13] if (len(row) > 13 and row[13] is not None) else 999
            areas.append((int(seq), row[0]))
    return [name for _, name in sorted(areas)]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 글로벌: 부하 분산 계산 (모든 배정 규칙에 공통 적용)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def calc_load_distribution(row: pd.Series) -> dict:
    """
    load_mh를 m_stdt~m_fndt 영업일에 균등 분배.
    반환: {monthly: {(yr,mo): mh}, weekly: {(iso_yr,iso_wk): mh}, daily_mh: float}
    """
    load_mh = row['load_mh']
    m_stdt  = row['m_stdt']
    m_fndt  = row['m_fndt']
    if pd.isna(load_mh) or pd.isna(m_stdt) or pd.isna(m_fndt):
        return {'monthly': {}, 'weekly': {}, 'daily_mh': 0.0}
    workdays = get_workdays_between(m_stdt.date(), m_fndt.date())
    if not workdays:
        return {'monthly': {}, 'weekly': {}, 'daily_mh': 0.0}
    daily_mh = load_mh / len(workdays)
    monthly, weekly = defaultdict(float), defaultdict(float)
    for d in workdays:
        monthly[(d.year, d.month)] += daily_mh
        iso = d.isocalendar()
        weekly[(iso.year, iso.week)] += daily_mh
    return {'monthly': dict(monthly), 'weekly': dict(weekly), 'daily_mh': daily_mh}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 글로벌 헬퍼: 부하 추적 & 배정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _init_area(area: str, cl: dict, cwl: dict):
    if area not in cl:
        cl[area]  = defaultdict(float)
        cwl[area] = defaultdict(float)


def add_load(area: str, dist: dict, cl: dict, cwl: dict):
    _init_area(area, cl, cwl)
    for k, v in dist['monthly'].items():
        cl[area][k] += v
    for k, v in dist['weekly'].items():
        cwl[area][k] += v


def remove_load(area: str, dist: dict, cl: dict, cwl: dict):
    for k, v in dist['monthly'].items():
        cl[area][k] -= v
    for k, v in dist['weekly'].items():
        cwl[area][k] -= v


def _check_3m_window(area_cl: dict, new_load: dict, area_tgt: dict) -> bool:
    """3개월 슬라이딩 윈도우 합산으로 목표 초과 여부 확인. True=배정 가능.
    단일 월이 목표를 넘어도 3개월 합산이 허용 범위면 배정 허용."""
    all_keys = sorted(set(new_load) | set(area_tgt))
    for i in range(len(all_keys)):
        window = all_keys[i:i + 3]
        total_load = sum(area_cl.get(k, 0.0) + new_load.get(k, 0.0) for k in window)
        total_tgt  = sum(area_tgt.get(k, 0)                          for k in window)
        if total_load > total_tgt:
            return False
    return True


def can_assign(area: str, dist: dict, cl: dict, tgt: dict) -> bool:
    """블록 1개 배정 시 3개월 슬라이딩 윈도우 합산 목표 초과 여부 확인. tgt에 없으면 무조건 True."""
    if area not in tgt:
        return True
    return _check_3m_window(cl.get(area, {}), dist['monthly'], tgt[area])


def can_assign_group(area: str, indices: list, load_dists: list,
                     cl: dict, tgt: dict) -> bool:
    """그룹 전체 배정 시 3개월 슬라이딩 윈도우 합산 목표 초과 여부 확인."""
    if area not in tgt:
        return True
    grp_load: dict = defaultdict(float)
    for idx in indices:
        for k, v in load_dists[idx]['monthly'].items():
            grp_load[k] += v
    return _check_3m_window(cl.get(area, {}), dict(grp_load), tgt[area])


def assign_block(df: pd.DataFrame, idx: int, area: str,
                 load_dists: list, cl: dict, cwl: dict):
    df.at[idx, 'assigned_area'] = area
    df.at[idx, 'rule_assigned']  = _current_rule
    add_load(area, load_dists[idx], cl, cwl)


def assign_roundrobin(candidates: list, areas: list,
                      df: pd.DataFrame, load_dists: list,
                      cl: dict, cwl: dict, tgt: dict) -> int:
    """
    candidates를 areas 순환(round-robin) 배정.
    특정 area가 현재 블록을 받지 못하면 그 블록만 건너뜀.
    모든 area가 특정 블록을 거부하면 해당 블록을 skip하고 다음 블록 시도.
    area가 연속으로 거부 횟수가 전체 후보 수를 초과하면 중단.
    반환: 배정 건수
    """
    area_idx    = 0
    cnt         = 0
    skip_streak = 0          # 연속 skip 횟수 (모든 area가 거부한 횟수)
    max_skip    = len(areas) # area 전체가 한 번씩 거부하면 streak 리셋
    for idx in candidates:
        assigned = False
        for offset in range(len(areas)):
            area = areas[(area_idx + offset) % len(areas)]
            if can_assign(area, load_dists[idx], cl, tgt):
                assign_block(df, idx, area, load_dists, cl, cwl)
                area_idx = (area_idx + offset + 1) % len(areas)
                cnt += 1
                skip_streak = 0
                assigned = True
                break
        if not assigned:
            skip_streak += 1
            if skip_streak >= max_skip * 3:
                break          # 연속 미배정이 지속되면 중단
    return cnt


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R0. 이관물량(H) 처리가능여부 리포팅
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r0_report_transfer_capacity(df: pd.DataFrame, load_dists: list, capa: dict):
    print("\n" + "=" * 70)
    print("  [R0] 이관물량(H) 처리가능여부 리포팅")
    print("=" * 70)
    year = 2026
    t_load: dict = defaultdict(float)
    h_load: dict = defaultdict(float)
    for i, row in df.iterrows():
        for (yr, mo), load in load_dists[i]['monthly'].items():
            if yr != year:
                continue
            if row['h_t'] == 'T':
                t_load[mo] += load
            else:
                h_load[mo] += load
    total_capa = {mo: sum(c.get(mo, 0) for a, c in capa.items() if a != AREA_Y9)
                  for mo in range(1, 13)}
    alert = False
    for mo in range(1, 13):
        t, h = t_load.get(mo, 0), h_load.get(mo, 0)
        total = t + h * 1.2
        cap   = total_capa.get(mo, 0)
        if total > cap:
            if not alert:
                print("  ⚠️  처리 초과 월 발생:")
                print(f"  {'월':^7} {'T부하':>10} {'H부하×1.2':>12} {'합계':>10} {'전체능력':>10} {'초과':>10}")
                print("  " + "-" * 63)
                alert = True
            print(f"  {year}-{mo:02d}  {t:>10,.0f} {h*1.2:>12,.0f} {total:>10,.0f} {cap:>10,.0f} {total-cap:>10,.0f}")
    if not alert:
        print("  ✅ 전체 월 처리 가능 (초과 없음)")
    print()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R1. 작업장별 월별 배정 목표물량 산출
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r1_calc_target_loads(df: pd.DataFrame, load_dists: list, capa: dict) -> dict:
    """
    월별 전체부하 / 전체능력 = 목표조업도
    목표조업도 × 작업장능력 = 작업장별 배정 목표물량
    반환: {area: {(year, month): target_mh}}
    """
    year = 2026
    total_load: dict = defaultdict(float)
    for dist in load_dists:
        for (yr, mo), load in dist['monthly'].items():
            if yr == year:
                total_load[(yr, mo)] += load
    total_capa = {(year, mo): sum(c.get(mo, 0) for a, c in capa.items() if a != AREA_Y9)
                  for mo in range(1, 13)}
    util = {k: (total_load[k] / total_capa[k] if total_capa[k] > 0 else 0)
            for k in total_capa}
    tgt: dict = {}
    for area, month_caps in capa.items():
        if area == AREA_Y9:
            continue
        tgt[area] = {(year, mo): month_caps.get(mo, 0) * util.get((year, mo), 0)
                     for mo in range(1, 13)}

    print("\n" + "=" * 90)
    print("  [R1] 작업장별 월별 배정 목표물량 (목표조업도 × 능력)")
    print("=" * 90)
    hdr = f"  {'작업장':^8}" + "".join(f"  {m:>2}월" for m in range(1, 13))
    print(hdr)
    print("  " + "-" * (len(hdr) - 2))
    for area in sorted(tgt):
        vals = [tgt[area].get((year, m), 0) for m in range(1, 13)]
        print(f"  {area:^8}" + "".join(f" {v:>6,.0f}" for v in vals))
    util_pct = [util.get((year, m), 0) * 100 for m in range(1, 13)]
    print(f"  {'목표조업도':^8}" + "".join(f" {u:>5.1f}%" for u in util_pct))
    print()
    return tgt


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R2. H/Y9 블록 → AREA15
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r2_assign_y9(df: pd.DataFrame, load_dists: list, cl: dict, cwl: dict):
    set_rule('R2')
    mask = (df['h_t'] == 'H') & df['blk'].str.startswith('Y9') & df['assigned_area'].isna()
    cnt = 0
    for idx in df[mask].index:
        assign_block(df, idx, AREA_Y9, load_dists, cl, cwl)
        cnt += 1
    print(f"  [R2] H/Y9 → AREA15: {cnt}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R3. T-중/대 → m_area
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r3_assign_t_mid_large(df: pd.DataFrame, load_dists: list, cl: dict, cwl: dict):
    set_rule('R3')
    mask = (df['h_t'] == 'T') & df['stg'].isin(['중', '대']) & df['assigned_area'].isna()
    cnt = 0
    for idx in df[mask].index:
        area = df.at[idx, 'm_area']
        if pd.notna(area) and area:
            assign_block(df, idx, area, load_dists, cl, cwl)
            cnt += 1
    print(f"  [R3] T-중/대 → m_area: {cnt}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R4. T-소/area_prior_1=동일 → prt_area
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r4_assign_t_small_dongil(df: pd.DataFrame, load_dists: list, cl: dict, cwl: dict):
    set_rule('R4')
    mask = (
        (df['h_t'] == 'T') & (df['stg'] == '소') &
        (df['area_prior_1'] == '동일') & df['assigned_area'].isna()
    )
    cnt = 0
    for idx in df[mask].index:
        area = df.at[idx, 'prt_area']
        if pd.notna(area) and area:
            assign_block(df, idx, area, load_dists, cl, cwl)
            cnt += 1
    print(f"  [R4] T-소/동일 → prt_area: {cnt}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R5. H-소 AREA3~5 처리가능여부 리포팅
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r5_report_h_small(df: pd.DataFrame, load_dists: list, cl: dict, capa: dict):
    print("\n" + "=" * 70)
    print("  [R5] H-소 물량 AREA3~5 처리가능여부 리포팅")
    print("=" * 70)
    year = 2026
    h_so_load: dict = defaultdict(float)
    mask = (df['h_t'] == 'H') & (df['stg'] == '소') & df['assigned_area'].isna()
    for i in df[mask].index:
        for (yr, mo), load in load_dists[i]['monthly'].items():
            if yr == year:
                h_so_load[mo] += load
    area35_remain: dict = defaultdict(float)
    for area in AREAS_SMALL:
        for mo in range(1, 13):
            cap = capa.get(area, {}).get(mo, 0)
            cur = cl.get(area, {}).get((year, mo), 0.0)
            area35_remain[mo] += max(0.0, cap - cur)
    alert = False
    for mo in range(1, 13):
        h   = h_so_load.get(mo, 0)
        rem = area35_remain.get(mo, 0)
        if h > rem:
            if not alert:
                print("  ⚠️  초과 월:")
                print(f"  {'월':^7} {'H-소부하':>12} {'AREA3~5잔여':>14} {'초과':>10}")
                print("  " + "-" * 47)
                alert = True
            print(f"  {year}-{mo:02d}  {h:>12,.0f} {rem:>14,.0f} {h-rem:>10,.0f}")
    if not alert:
        print("  ✅ AREA3~5에서 H-소 전량 처리 가능")
    print()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R6. H-소 → AREA3~5 순환(round-robin)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r6_assign_h_small(df: pd.DataFrame, load_dists: list,
                      cl: dict, cwl: dict, tgt: dict):
    set_rule('R6')
    mask  = (df['h_t'] == 'H') & (df['stg'] == '소') & df['assigned_area'].isna()
    cands = df[mask].sort_values('m_stdt').index.tolist()
    cnt   = assign_roundrobin(cands, AREAS_SMALL, df, load_dists, cl, cwl, tgt)
    print(f"  [R6] H-소 → AREA3~5 순환: {cnt}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R7. H-중-jig_F → AREA2
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r7_assign_h_mid_jig_f(df: pd.DataFrame, load_dists: list,
                           cl: dict, cwl: dict, tgt: dict):
    set_rule('R7')
    area = 'AREA2'
    SPECIAL_BLK = {'D114', 'D174', 'D204'}

    # Part1: ship_kind=A + 특정 blk 앞4자리 → 목표물량 무관 무조건 AREA2
    mask_p1 = (
        (df['h_t'] == 'H') & (df['ship_kind'] == 'A') & (df['stg'] == '중') &
        df['blk'].str[:4].isin(SPECIAL_BLK) & df['assigned_area'].isna()
    )
    cnt_p1 = 0
    for idx in df[mask_p1].index:
        assign_block(df, idx, area, load_dists, cl, cwl)
        cnt_p1 += 1

    # Part2: H-중/jig=F → AREA2 목표물량까지 (이미 배정된 블록 제외)
    # 특정 월이 목표에 도달해 거부된 블록은 건너뛰고 계속 시도
    mask_p2 = (
        (df['h_t'] == 'H') & (df['stg'] == '중') &
        (df['jig'] == 'F') & df['assigned_area'].isna()
    )
    cnt_p2 = 0
    skip_streak = 0
    for idx in df[mask_p2].sort_values('m_stdt').index:
        if can_assign(area, load_dists[idx], cl, tgt):
            assign_block(df, idx, area, load_dists, cl, cwl)
            cnt_p2 += 1
            skip_streak = 0
        else:
            skip_streak += 1
            if skip_streak >= 50:   # 연속 50건 거부 시 AREA2 용량 소진으로 판단
                break
    print(f"  [R7] H-중-jig_F → AREA2: Part1={cnt_p1}건 + Part2={cnt_p2}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R8. H-중-jig_D/L → AREA1→2→6 순환
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r8_assign_h_mid_jig_dl(df: pd.DataFrame, load_dists: list,
                            cl: dict, cwl: dict, tgt: dict):
    set_rule('R8')
    mask = (
        (df['h_t'] == 'H') & (df['stg'] == '중') &
        ~df['blk'].str.startswith('H') &
        ~df['blk'].str[:3].isin(['E11', 'E51']) &
        df['jig'].isin(['D', 'L', 'W']) & df['assigned_area'].isna()
    )
    cands = df[mask].sort_values('m_stdt').index.tolist()
    cnt   = assign_roundrobin(cands, AREAS_MID_DL, df, load_dists, cl, cwl, tgt)
    print(f"  [R8] H-중-jig_D/L/W → AREA1→2→6 순환: {cnt}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R9. H-전문화 family 배정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r9_assign_specialized_family(df: pd.DataFrame, load_dists: list,
                                  cl: dict, cwl: dict, tgt: dict):
    set_rule('R9')
    def _assign(blk_filter, suffix: str, area: str, label: str):
        base = (df['h_t'] == 'H') & df['stg'].isin(['대', '중']) & df['assigned_area'].isna()
        sub  = df[base & blk_filter & (df['blk'].str[-1] == suffix)].copy()
        sub['_grp'] = sub['pjt'].astype(str) + '|' + sub['blk'].str[:3]
        # 그룹 착수일 오름차순
        grp_order = sub.groupby('_grp')['m_stdt'].min().sort_values().index.tolist()
        cnt = 0
        skip_streak = 0
        for grp in grp_order:
            g_idx = sub[sub['_grp'] == grp].index.tolist()
            for idx in df.loc[g_idx].sort_values('m_stdt').index:
                if df.at[idx, 'assigned_area'] is not None:
                    continue
                if can_assign(area, load_dists[idx], cl, tgt):
                    assign_block(df, idx, area, load_dists, cl, cwl)
                    cnt += 1
                    skip_streak = 0
                else:
                    skip_streak += 1
                    if skip_streak >= 50:
                        print(f"  [R9-{label}] → {area}: {cnt}건")
                        return
        print(f"  [R9-{label}] → {area}: {cnt}건")

    # 1-1: blk 첫째자리 H, 마지막자리 P → AREA7
    _assign(df['blk'].str.startswith('H'), 'P', 'AREA7',  '1-1')
    # 1-2: blk 첫째자리 H, 마지막자리 S → AREA8
    _assign(df['blk'].str.startswith('H'), 'S', 'AREA8',  '1-2')
    # 2-1: blk E11 또는 F51 시작, 마지막자리 P → AREA9
    _assign(df['blk'].str[:3].isin(['E11', 'F51']), 'P', 'AREA9',  '2-1')
    # 2-2: blk E11 또는 F51 시작, 마지막자리 S → AREA10
    _assign(df['blk'].str[:3].isin(['E11', 'F51']), 'S', 'AREA10', '2-2')


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R10. H-대-jig_D/L 그룹 → AREA12
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r10_assign_h_large_jig_dl(df: pd.DataFrame, load_dists: list,
                               cl: dict, cwl: dict, tgt: dict):
    set_rule('R10')
    area = 'AREA12'
    seed_mask = (
        (df['h_t'] == 'H') & (df['stg'] == '대') &
        df['jig'].isin(['D', 'L']) & df['assigned_area'].isna()
    )
    # 씨드 블록 그룹키 수집: (pjt, blk[0:3], blk[-1])
    grp_keys = set(
        df[seed_mask].apply(lambda r: (r['pjt'], r['blk'][:3], r['blk'][-1]), axis=1)
    )

    def in_grp(r):
        return (r['pjt'], r['blk'][:3], r['blk'][-1]) in grp_keys

    all_h_unassigned = (df['h_t'] == 'H') & df['assigned_area'].isna()
    cands = (df[all_h_unassigned][df[all_h_unassigned].apply(in_grp, axis=1)]
             .sort_values('m_stdt').index.tolist())
    cnt = 0
    skip_streak = 0
    for idx in cands:
        if can_assign(area, load_dists[idx], cl, tgt):
            assign_block(df, idx, area, load_dists, cl, cwl)
            cnt += 1
            skip_streak = 0
        else:
            skip_streak += 1
            if skip_streak >= 50:
                break
    print(f"  [R10] H-대-jig_D/L 그룹 → AREA12: {cnt}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R11. H-대/중 잔여물량 배정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r11_assign_h_remaining(df: pd.DataFrame, load_dists: list,
                            cl: dict, cwl: dict, tgt: dict):
    set_rule('R11')
    mask = (df['h_t'] == 'H') & df['stg'].isin(['대', '중']) & df['assigned_area'].isna()
    sub  = df[mask].copy()
    sub['_grp_key'] = sub['pjt'].astype(str) + '|' + sub['blk'].str[:3] + '|' + sub['blk'].str[-1]
    sub['_ps_key']  = sub['pjt'].astype(str) + '|' + sub['blk'].str[:3]

    # PS 페어 감지: 동일 (pjt, blk[0:3])에 P와 S 접미사 그룹 모두 존재
    sfx_by_ps = sub.groupby('_ps_key')['_grp_key'].apply(lambda ks: {k.split('|')[2] for k in ks})
    ps_pairs  = {k for k, s in sfx_by_ps.items() if 'P' in s and 'S' in s}

    # 그룹 착수일 기준 처리 순서
    grp_min_stdt = sub.groupby('_grp_key')['m_stdt'].min().sort_values()
    processed: set = set()
    cnt = 0

    def assign_individually(idx_list: list):
        nonlocal cnt
        for i in sorted(idx_list, key=lambda x: df.at[x, 'm_stdt']):
            if df.at[i, 'assigned_area'] is not None:
                continue
            for a in AREAS_REMAIN:
                if can_assign(a, load_dists[i], cl, tgt):
                    assign_block(df, i, a, load_dists, cl, cwl)
                    cnt += 1
                    break

    for grp_key in grp_min_stdt.index:
        if grp_key in processed:
            continue
        parts = grp_key.split('|')
        pjt_v, blk3, sfx = int(parts[0]), parts[1], parts[2]
        ps_key    = f"{pjt_v}|{blk3}"
        unassigned = [i for i in sub[sub['_grp_key'] == grp_key].index
                      if df.at[i, 'assigned_area'] is None]
        if not unassigned:
            processed.add(grp_key)
            continue

        # ── PS 페어 배정 ─────────────────────────────────────────────────────
        if ps_key in ps_pairs and sfx in ('P', 'S'):
            partner_sfx = 'S' if sfx == 'P' else 'P'
            partner_key = f"{pjt_v}|{blk3}|{partner_sfx}"
            processed.add(grp_key)
            processed.add(partner_key)

            p_key = grp_key     if sfx == 'P' else partner_key
            s_key = partner_key if sfx == 'P' else grp_key
            p_idx = [i for i in sub[sub['_grp_key'] == p_key].index
                     if df.at[i, 'assigned_area'] is None]
            s_idx = [i for i in sub[sub['_grp_key'] == s_key].index
                     if df.at[i, 'assigned_area'] is None]

            pair_done = False
            for (p_area, s_area) in PS_AREA_PAIRS:
                if (can_assign_group(p_area, p_idx, load_dists, cl, tgt) and
                        can_assign_group(s_area, s_idx, load_dists, cl, tgt)):
                    for i in p_idx:
                        assign_block(df, i, p_area, load_dists, cl, cwl)
                        cnt += 1
                    for i in s_idx:
                        assign_block(df, i, s_area, load_dists, cl, cwl)
                        cnt += 1
                    pair_done = True
                    break
            if not pair_done:
                assign_individually(p_idx + s_idx)

        # ── 단독 그룹 배정 ───────────────────────────────────────────────────
        else:
            processed.add(grp_key)
            grp_done = False
            for a in AREAS_REMAIN:
                if can_assign_group(a, unassigned, load_dists, cl, tgt):
                    for i in unassigned:
                        assign_block(df, i, a, load_dists, cl, cwl)
                        cnt += 1
                    grp_done = True
                    break
            if not grp_done:
                assign_individually(unassigned)

    print(f"  [R11] H-대/중 잔여 → AREA7~14: {cnt}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R12. H-dir 블록 → prt_area (override)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r12_assign_h_dir(df: pd.DataFrame, load_dists: list, cl: dict, cwl: dict):
    set_rule('R12')
    mask = (df['h_t'] == 'H') & df['blk'].str[3].isin(['9', 'H'])
    cnt = 0
    for idx in df[mask].index:
        area = df.at[idx, 'prt_area']
        if not (pd.notna(area) and area):
            continue  # prt_area 공란 → skip
        old = df.at[idx, 'assigned_area']
        if old and old != area:
            remove_load(old, load_dists[idx], cl, cwl)
        assign_block(df, idx, area, load_dists, cl, cwl)
        cnt += 1
    print(f"  [R12] H-dir(blk[3]=9/H) → prt_area override: {cnt}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# R13. T-소 분산배정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

AREA_PRIOR_COLS = ['area_prior_1', 'area_prior_2', 'area_prior_3',
                   'area_prior_4', 'area_prior_5']

def r13_assign_t_small_dist(df: pd.DataFrame, load_dists: list,
                             cl: dict, cwl: dict, tgt: dict,
                             area_seq: list):
    """
    T-소 미배정 블록을 area_seq 순서로 분산 배정.
    - (pjt, blk, type) 그룹핑
    - 그룹 정렬: assign_prior 오름차순(낮은수=높은우선순위) → min(m_stdt) 오름차순
    - area별: area_prior_1 매칭 우선 → 미달 시 area_prior_2~5 순차 추가
    - 목표물량 도달 시 해당 area 중단, 다음 area 진행
    """
    set_rule('R13')
    total_cnt = 0

    for area in area_seq:
        if area not in tgt:
            continue
        area_cnt = 0

        for prior_col in AREA_PRIOR_COLS:
            # 현재 prior_col에서 area와 매칭되는 미배정 T-소 블록
            sub_mask = (
                (df['h_t'] == 'T') & (df['stg'] == '소') &
                df['assigned_area'].isna() & (df[prior_col] == area)
            )
            sub = df[sub_mask].copy()
            if sub.empty:
                continue

            # (pjt, blk, type) 그룹키
            sub['_grp'] = (sub['pjt'].astype(str) + '|' +
                           sub['blk'].astype(str) + '|' +
                           sub['type'].astype(str))

            # 그룹별 대표 assign_prior(최솟값)와 최소 m_stdt로 그룹 순서 결정
            grp_info = (sub.groupby('_grp')
                        .agg(ap=('assign_prior', 'min'), ms=('m_stdt', 'min'))
                        .sort_values(['ap', 'ms']))
            grp_order = grp_info.index.tolist()

            area_skip = 0   # 이 prior_col 내 연속 거부 횟수

            for grp_key in grp_order:
                grp_idx = sub[sub['_grp'] == grp_key].sort_values('m_stdt').index.tolist()

                for idx in grp_idx:
                    if df.at[idx, 'assigned_area'] is not None:
                        continue
                    if can_assign(area, load_dists[idx], cl, tgt):
                        assign_block(df, idx, area, load_dists, cl, cwl)
                        area_cnt += 1
                        total_cnt += 1
                        area_skip = 0
                    else:
                        area_skip += 1

                if area_skip >= 50:
                    break   # 이 prior_col 내 연속 거부 50건 → 다음 prior_col 시도

        print(f"    [R13] {area}: {area_cnt}건")

    print(f"  [R13] T-소 분산배정 합계: {total_cnt}건")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 결과 요약 출력 & 저장
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def print_utilization_summary(df: pd.DataFrame, cl: dict, capa: dict):
    year = 2026
    print("\n" + "=" * 90)
    print("  배정 결과 — 작업장별 월별 조업도 (배정부하 / 능력 × 100%)")
    print("=" * 90)
    hdr = f"  {'작업장':^8}" + "".join(f"  {m:>2}월" for m in range(1, 13)) + "  연평균"
    print(hdr)
    print("  " + "-" * (len(hdr) - 2))
    all_utils = []
    for area in sorted(a for a in capa if a != AREA_Y9):
        utils = []
        for m in range(1, 13):
            cap  = capa[area].get(m, 0)
            load = cl.get(area, {}).get((year, m), 0.0)
            utils.append((load / cap * 100) if cap > 0 else 0.0)
        all_utils.append(utils)
        avg = sum(utils) / len(utils)
        print(f"  {area:^8}" + "".join(f" {u:>5.1f}%" for u in utils) + f" {avg:>6.1f}%")
    print("  " + "-" * (len(hdr) - 2))
    col_avg = [sum(r[i] for r in all_utils) / max(len(all_utils), 1) for i in range(12)]
    print(f"  {'전체평균':^8}" + "".join(f" {u:>5.1f}%" for u in col_avg) + f" {sum(col_avg)/12:>6.1f}%")

    total = len(df)
    asgn  = df['assigned_area'].notna().sum()
    print(f"\n  전체: {total:,}건 | 배정완료: {asgn:,}건 | 미배정: {total-asgn:,}건")
    print("\n  [작업장별 배정 건수]")
    for area, c in df['assigned_area'].value_counts().sort_index().items():
        print(f"    {area}: {c:,}건")


RULE_LABELS = {
    'R2':  'H/Y9 → AREA15',
    'R3':  'T-중/대 → m_area',
    'R4':  'T-소/동일 → prt_area',
    'R6':  'H-소 → AREA3~5 순환',
    'R7':  'H-중-jig_F → AREA2',
    'R8':  'H-중-jig_D/L → AREA1→2→6',
    'R9':  'H-전문화 Family',
    'R10': 'H-대-jig_D/L 그룹 → AREA12',
    'R11': 'H-대/중 잔여 → AREA7~14',
    'R12': 'H-dir → prt_area(override)',
    'R13': 'T-소 분산배정 (area_seq 순)',
}

def save_stats(df: pd.DataFrame, cl: dict, capa: dict, tgt: dict):
    year = 2026
    # 단계별 배정 건수 & 작업장 분포
    steps = []
    for rule in ['R2','R3','R4','R6','R7','R8','R9','R10','R11','R12','R13']:
        sub = df[df['rule_assigned'] == rule]
        area_dist = sub['assigned_area'].value_counts().to_dict()
        steps.append({
            'rule':   rule,
            'label':  RULE_LABELS.get(rule, rule),
            'count':  int(len(sub)),
            'areas':  {k: int(v) for k, v in area_dist.items()},
        })

    # 작업장별 월별 조업도
    utilization = {}
    for area in sorted(a for a in capa if a != AREA_Y9):
        utilization[area] = {}
        for mo in range(1, 13):
            cap  = capa[area].get(mo, 0)
            load = cl.get(area, {}).get((year, mo), 0.0)
            util = round(load / cap * 100, 1) if cap > 0 else 0.0
            utilization[area][f'{year}-{mo:02d}'] = util

    # 작업장별 목표 조업도
    target_util = {}
    for area in sorted(a for a in capa if a != AREA_Y9):
        target_util[area] = {}
        for mo in range(1, 13):
            cap = capa[area].get(mo, 0)
            t   = tgt.get(area, {}).get((year, mo), 0)
            target_util[area][f'{year}-{mo:02d}'] = round(t / cap * 100, 1) if cap > 0 else 0.0

    stats = {
        'total':       int(len(df)),
        'assigned':    int(df['assigned_area'].notna().sum()),
        'unassigned':  int(df['assigned_area'].isna().sum()),
        'steps':       steps,
        'utilization': utilization,
        'target_util': target_util,
    }
    with open(STATS_PATH, 'w', encoding='utf-8') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    print(f"📊 통계 저장 완료: {STATS_PATH}")
    return stats


def save_result(df: pd.DataFrame, path: str):
    out = df.copy()
    out['m_stdt'] = out['m_stdt'].dt.date
    out['m_fndt'] = out['m_fndt'].dt.date
    out.to_excel(path, index=False)
    print(f"\n✅ 저장 완료: {path}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HTML 리포트 생성
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

MERMAID_CHART = """
flowchart TD
    START([▶ 시작]) --> LOAD[📂 데이터 로드<br/>blk_master + area_capa]
    LOAD --> CALC[⚙️ 글로벌: load_mh 부하 분산<br/>m_stdt~m_fndt 영업일 균등 배분<br/>→ 월별 부하 + 주별 부하 집계]

    CALC --> R0{📊 R0 이관물량 처리가능 리포팅<br/>T부하 + H부하×1.2 vs 전체능력<br/>초과 시 월별 경고 출력}
    R0 --> R1[📌 R1 목표물량 산출<br/>목표조업도 = 전체부하 ÷ 전체능력<br/>배정목표 = 목표조업도 × 작업장능력]

    R1 --> R2[R2 H행 + blk Y9시작<br/>→ AREA15 무조건 배정]
    R2 --> R3[R3 T행 + stg=중·대<br/>→ m_area 배정]
    R3 --> R4[R4 T행 + stg=소 + area_prior_1=동일<br/>→ prt_area 배정]

    R4 --> R5{📊 R5 H-소 처리가능 리포팅<br/>H-소 부하 vs AREA3~5 잔여능력<br/>초과 시 월별 경고 출력}
    R5 --> R6[R6 H행 + stg=소<br/>AREA3 → AREA4 → AREA5 순환 배정<br/>목표물량 도달 시 해당 area 중단]

    R6 --> R7[R7-Part1 H+중+ship_kind=A+blk∈D114·D174·D204<br/>→ AREA2 무조건 배정<br/>R7-Part2 H+중+jig=F<br/>→ AREA2 목표물량까지 배정]

    R7 --> R8[R8 H행 + stg=중 + jig=D·L<br/>단 H·E11·E51 시작 블록 제외<br/>AREA1 → AREA2 → AREA6 순환 배정]

    R8 --> R9[R9 H행 + stg=대·중 전문화 Family<br/>1-1: blk H시작 + 말미P → AREA7<br/>1-2: blk H시작 + 말미S → AREA8<br/>2-1: blk E11·F51시작 + 말미P → AREA9<br/>2-2: blk E11·F51시작 + 말미S → AREA10]

    R9 --> R10[R10 H행 + stg=대 + jig=D·L<br/>동일 그룹 포함하여 → AREA12<br/>목표물량 도달 시 중단]

    R10 --> R11[R11 H행 + stg=대·중 잔여물량<br/>PS페어 그룹: AREA7+8 · AREA9+10 · AREA13+14<br/>단독 그룹: AREA7→14 순차 first-fit<br/>그룹 초과 시 해제 후 개별 배정]

    R11 --> R12[R12 H행 + blk 4번째 자리 9 또는 H<br/>→ prt_area로 override 배정]

    R12 --> R13[R13 T행 + stg=소 미배정 잔여<br/>area_seq 순으로 분산 배정<br/>그룹: pjt+blk+type / 정렬: assign_prior→m_stdt<br/>area_prior_1→5 단계적 확장 / 목표 도달 시 다음 area]
    R13 --> SAVE[💾 blk_assign_result.xlsx 저장]

    style R0 fill:#fff3cd,stroke:#ffc107
    style R5 fill:#fff3cd,stroke:#ffc107
    style CALC fill:#d1ecf1,stroke:#0c7cd5
    style R1 fill:#d4edda,stroke:#28a745
    style SAVE fill:#d4edda,stroke:#28a745
"""


def _util_color(v: float) -> str:
    if v == 0:
        return '#f8f9fa'
    elif v < 60:
        return f'hsl(210,70%,{max(75, 95 - v * 0.3):.0f}%)'
    elif v < 90:
        return f'hsl(120,60%,{max(55, 85 - (v-60)*0.7):.0f}%)'
    elif v < 110:
        return f'hsl(45,90%,{max(60, 80 - (v-90)*.5):.0f}%)'
    else:
        return f'hsl(0,80%,{max(50, 75 - (v-110)*.4):.0f}%)'


def generate_html_report(df: pd.DataFrame, stats: dict, agent_run_time: datetime):
    agent_run_time_str = agent_run_time.strftime('%Y-%m-%d %H:%M:%S')

    dfw = df.copy()
    dfw['m_stdt'] = pd.to_datetime(dfw['m_stdt']).dt.strftime('%Y-%m-%d')
    dfw['m_fndt'] = pd.to_datetime(dfw['m_fndt']).dt.strftime('%Y-%m-%d')
    dfw['assigned_area'] = dfw['assigned_area'].fillna('')
    dfw['rule_assigned'] = dfw['rule_assigned'].fillna('')
    dfw['prt_area']      = dfw['prt_area'].fillna('')

    block_cols  = ['pjt', 'h_t', 'blk', 'stg', 'm_stdt', 'm_fndt',
                   'm_area', 'load_mh', 'assigned_area', 'rule_assigned']
    blocks_json = dfw[block_cols].fillna('').to_dict(orient='records')

    months = [f'2026-{m:02d}' for m in range(1, 13)]
    util   = stats['utilization']
    tgt_u  = stats['target_util']

    util_rows = ''
    for area in sorted(util.keys()):
        vals     = [util[area].get(m, 0) for m in months]
        avg      = sum(vals) / len(vals)
        tgt_vals = [tgt_u.get(area, {}).get(m, 0) for m in months]
        cells = ''
        for i, v in enumerate(vals):
            bg = _util_color(v)
            tv = tgt_vals[i]
            cells += (f'<td style="background:{bg};text-align:center;font-size:12px"'
                      f' title="목표:{tv:.1f}%">{v:.1f}%</td>')
        avg_bg = _util_color(avg)
        util_rows += (f'<tr><td class="fw-semibold">{area}</td>{cells}'
                      f'<td style="background:{avg_bg};text-align:center;font-size:12px">'
                      f'{avg:.1f}%</td></tr>\n')

    step_rows  = ''
    cumulative = 0
    for s in stats['steps']:
        cumulative += s['count']
        areas_str = ', '.join(f"{a}:{c}" for a, c in sorted(s['areas'].items()))
        bar_width = min(100, s['count'] * 2) if s['count'] > 0 else 0
        step_rows += f"""<tr>
        <td><span class="badge bg-primary">{s['rule']}</span></td>
        <td>{s['label']}</td>
        <td class="text-end fw-bold">{s['count']:,}</td>
        <td class="text-end text-muted">{cumulative:,}</td>
        <td><div style="background:#0d6efd;height:16px;width:{bar_width}px;border-radius:3px;display:inline-block"></div></td>
        <td style="font-size:11px">{areas_str or '-'}</td>
    </tr>\n"""

    MONTHS_HDR = ''.join(f'<th>{m[5:]}</th>' for m in months)

    HTML = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>블록 작업장 배정 Dashboard</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script type="module">
  import mermaid from 'https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.esm.min.mjs';
  mermaid.initialize({{ startOnLoad:true, theme:'default', flowchart:{{curve:'basis'}} }});
</script>
<style>
body{{ font-family:'Segoe UI',sans-serif; background:#f5f7fa; }}
.nav-tabs .nav-link{{ color:#495057; font-weight:500; }}
.nav-tabs .nav-link.active{{ color:#0d6efd; border-bottom:3px solid #0d6efd; }}
.card{{ border:none; border-radius:12px; box-shadow:0 2px 8px rgba(0,0,0,.08); }}
.stat-box{{ text-align:center; padding:20px; border-radius:10px; }}
.mermaid{{ max-width:100%; overflow-x:auto; }}
#blockTable{{ font-size:13px; }}
#blockTable td,#blockTable th{{ padding:5px 8px; }}
.filter-bar {{ background:#fff; padding:12px 16px; border-radius:8px; margin-bottom:12px; }}
.legend-item{{ display:inline-block; width:16px; height:16px; border-radius:3px; margin-right:4px; vertical-align:middle; }}
.ts-bar{{ font-size:11px; color:#adb5bd; }}
</style>
</head>
<body>

<nav class="navbar navbar-dark bg-dark px-4 py-2">
  <span class="navbar-brand fw-bold">🏗️ 블록 작업장 배정 Dashboard</span>
  <div class="d-flex flex-column align-items-end gap-1">
    <span class="text-light small">전체: {stats['total']:,}건 | 배정완료: {stats['assigned']:,}건 | 미배정: {stats['unassigned']:,}건</span>
    <span class="ts-bar">
      ⚙️ Agent 실행: {agent_run_time_str}
      &nbsp;|&nbsp;
      🌐 페이지 로드: <span id="pageLoadTime">-</span>
    </span>
  </div>
</nav>

<div class="container-fluid py-3 px-4">

  <!-- KPI 카드 -->
  <div class="row g-3 mb-3">
    <div class="col-md-3">
      <div class="card stat-box h-100">
        <div class="text-muted small">전체 블록</div>
        <div class="fs-2 fw-bold text-dark">{stats['total']:,}</div>
      </div>
    </div>
    <div class="col-md-3">
      <div class="card stat-box h-100">
        <div class="text-muted small">배정 완료</div>
        <div class="fs-2 fw-bold text-success">{stats['assigned']:,}</div>
        <div class="text-muted small">{stats['assigned']/stats['total']*100:.1f}%</div>
      </div>
    </div>
    <div class="col-md-3">
      <div class="card stat-box h-100">
        <div class="text-muted small">미배정</div>
        <div class="fs-2 fw-bold text-warning">{stats['unassigned']:,}</div>
        <div class="text-muted small">{stats['unassigned']/stats['total']*100:.1f}%</div>
      </div>
    </div>
    <div class="col-md-3">
      <div class="card stat-box h-100">
        <div class="text-muted small">배정 규칙 수</div>
        <div class="fs-2 fw-bold text-primary">{len(stats['steps'])}</div>
        <div class="text-muted small">R2 ~ R13</div>
      </div>
    </div>
  </div>

  <!-- 탭 네비게이션 -->
  <ul class="nav nav-tabs mb-3" id="mainTab">
    <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#tab-flow">🔄 배정 플로우챠트</button></li>
    <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tab-steps">📊 단계별 배정 결과</button></li>
    <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tab-util">🏭 작업장별 조업도</button></li>
    <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tab-blocks">📋 블록 목록</button></li>
  </ul>

  <div class="tab-content">

    <!-- ───────────── Tab 1: 플로우챠트 ───────────── -->
    <div class="tab-pane fade show active" id="tab-flow">
      <div class="card p-4">
        <h5 class="mb-3">📌 배정 규칙 전체 플로우챠트</h5>
        <div class="alert alert-info small mb-3">
          🟡 노란 박스: 리포팅 단계 (배정 없음) &nbsp;|&nbsp;
          🔵 파란 박스: 글로벌 계산 단계 &nbsp;|&nbsp;
          🟢 초록 박스: 목표 산출 및 저장
        </div>
        <div class="mermaid" style="min-height:600px">
{MERMAID_CHART}
        </div>
        <hr class="my-4">
        <h6>규칙별 요약</h6>
        <div class="row g-2">
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R0</b> 이관물량 처리가능여부: T부하 + H부하×1.2 vs 전체능력 비교 → 초과 월 경고</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R1</b> 목표물량 산출: 전체조업도 = Σ부하 ÷ Σ능력 → 각 작업장 배정목표 = 조업도 × 능력</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R2</b> H/Y9 블록: blk이 Y9로 시작하는 H행 → AREA15 (목표 무관)</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R3</b> T-중/대: h_t=T이고 stg=중 또는 대 → m_area 값으로 직접 배정</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R4</b> T-소/동일: h_t=T, stg=소, area_prior_1=동일 → prt_area 배정</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R5</b> H-소 처리가능여부: H-소 부하 vs AREA3~5 잔여능력 비교 → 초과 월 경고</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R6</b> H-소: m_stdt 오름차순 정렬 후 AREA3→4→5 순환. 목표 도달 area 건너뜀</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R7</b> H-중-jig_F: Part1=특정블록 무조건 AREA2 / Part2=나머지 jig_F 목표까지</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R8</b> H-중-jig_D/L: H·E11·E51 시작 제외. AREA1→2→6 순환 배정</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R9</b> 전문화 Family: (pjt+blk앞3자리)로 그룹. P말미→AREA7/9, S말미→AREA8/10</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R10</b> H-대-jig_D/L: 씨드블록과 동일 (pjt+blk앞3자리+말미) 그룹 → AREA12</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R11</b> 잔여: PS페어→(7+8)(9+10)(13+14) / 단독→AREA7~14 first-fit / 초과시 개별</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R12</b> H-dir: blk 4번째 자리가 9 또는 H → prt_area로 강제 override 배정</div></div>
          <div class="col-md-6"><div class="p-2 border rounded small"><b>R13</b> T-소 분산배정: area_seq 순, (pjt+blk+type) 그룹, assign_prior→m_stdt 정렬. area_prior_1~5 단계적 확장, 목표 도달 시 다음 area</div></div>
        </div>
      </div>
    </div>

    <!-- ───────────── Tab 2: 단계별 결과 ───────────── -->
    <div class="tab-pane fade" id="tab-steps">
      <div class="row g-3">
        <div class="col-lg-5">
          <div class="card p-3 h-100">
            <h6 class="mb-3">단계별 배정 건수</h6>
            <canvas id="stepChart" height="320"></canvas>
          </div>
        </div>
        <div class="col-lg-7">
          <div class="card p-3 h-100">
            <h6 class="mb-3">단계별 배정 상세</h6>
            <table class="table table-sm table-hover mb-0">
              <thead class="table-light">
                <tr>
                  <th>규칙</th><th>설명</th><th class="text-end">배정</th>
                  <th class="text-end">누적</th><th>비율</th><th>배정 작업장</th>
                </tr>
              </thead>
              <tbody>{step_rows}</tbody>
            </table>
          </div>
        </div>
      </div>

      <div class="card mt-3 p-3">
        <h6 class="mb-3">작업장별 배정 건수</h6>
        <canvas id="areaChart" height="120"></canvas>
      </div>
    </div>

    <!-- ───────────── Tab 3: 조업도 ───────────── -->
    <div class="tab-pane fade" id="tab-util">
      <div class="card p-3">
        <div class="d-flex justify-content-between align-items-center mb-2">
          <h6 class="mb-0">작업장별 월별 조업도 (배정부하 / 능력 × 100%)</h6>
          <div class="small">
            <span class="legend-item" style="background:#d1ecf1"></span>0~60%&nbsp;
            <span class="legend-item" style="background:#78c98a"></span>60~90%&nbsp;
            <span class="legend-item" style="background:#ffc107"></span>90~110%&nbsp;
            <span class="legend-item" style="background:#dc3545"></span>110%+
          </div>
        </div>
        <div class="table-responsive">
          <table class="table table-sm table-bordered mb-2" style="font-size:12px">
            <thead class="table-dark">
              <tr><th>작업장</th>{MONTHS_HDR}<th>연평균</th></tr>
            </thead>
            <tbody>{util_rows}</tbody>
          </table>
        </div>
        <div class="small text-muted">※ 셀에 마우스를 올리면 목표 조업도 확인 가능</div>
        <div class="mt-3">
          <canvas id="utilChart" height="120"></canvas>
        </div>
      </div>
    </div>

    <!-- ───────────── Tab 4: 블록 목록 ───────────── -->
    <div class="tab-pane fade" id="tab-blocks">
      <div class="filter-bar d-flex gap-3 flex-wrap align-items-center">
        <div>
          <label class="form-label mb-0 small fw-semibold">H/T</label>
          <select id="fHT" class="form-select form-select-sm" style="width:90px">
            <option value="">전체</option>
            <option>H</option><option>T</option>
          </select>
        </div>
        <div>
          <label class="form-label mb-0 small fw-semibold">규격</label>
          <select id="fStg" class="form-select form-select-sm" style="width:90px">
            <option value="">전체</option>
            <option>대</option><option>중</option><option>소</option>
          </select>
        </div>
        <div>
          <label class="form-label mb-0 small fw-semibold">배정규칙</label>
          <select id="fRule" class="form-select form-select-sm" style="width:100px">
            <option value="">전체</option>
            <option>R2</option><option>R3</option><option>R4</option>
            <option>R6</option><option>R7</option><option>R8</option>
            <option>R9</option><option>R10</option><option>R11</option><option>R12</option>
            <option>R13</option>
            <option value="NONE">미배정</option>
          </select>
        </div>
        <div>
          <label class="form-label mb-0 small fw-semibold">작업장</label>
          <select id="fArea" class="form-select form-select-sm" style="width:120px">
            <option value="">전체</option>
          </select>
        </div>
        <div class="ms-auto">
          <input id="fSearch" type="text" class="form-control form-control-sm" placeholder="🔍 blk 검색..." style="width:160px">
        </div>
        <div><span id="rowCount" class="badge bg-secondary">0건</span></div>
      </div>

      <div class="card p-0">
        <div style="max-height:520px;overflow-y:auto">
          <table class="table table-sm table-hover mb-0" id="blockTable">
            <thead class="table-dark sticky-top">
              <tr>
                <th>pjt</th><th>h_t</th><th>blk</th><th>stg</th>
                <th>m_stdt</th><th>m_fndt</th><th>m_area</th>
                <th>load_mh</th><th>assigned_area</th><th>rule</th>
              </tr>
            </thead>
            <tbody id="blockBody"></tbody>
          </table>
        </div>
      </div>
    </div>

  </div><!-- tab-content -->
</div><!-- container -->

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
<script>
// ── 페이지 로드 시점 표시 ────────────────────────────────────────────────────
document.getElementById('pageLoadTime').textContent =
  new Date().toLocaleString('ko-KR', {{
    year:'numeric', month:'2-digit', day:'2-digit',
    hour:'2-digit', minute:'2-digit', second:'2-digit', hour12:false
  }});

// ── 임베디드 데이터 ──────────────────────────────────────────────────────────
const STATS  = {json.dumps(stats, ensure_ascii=False)};
const BLOCKS = {json.dumps(blocks_json, ensure_ascii=False)};

// ── Chart: 단계별 배정 건수 ─────────────────────────────────────────────────
const stepLabels = STATS.steps.map(s => s.rule);
const stepData   = STATS.steps.map(s => s.count);
new Chart(document.getElementById('stepChart'), {{
  type: 'bar',
  data: {{
    labels: stepLabels,
    datasets: [{{ label: '배정 건수', data: stepData,
      backgroundColor: stepData.map(v =>
        v > 100 ? '#0d6efd' : v > 50 ? '#198754' : v > 10 ? '#fd7e14' : '#6c757d'
      ), borderRadius: 4
    }}]
  }},
  options: {{ plugins:{{legend:{{display:false}}}}, scales:{{y:{{beginAtZero:true}}}} }}
}});

// ── Chart: 작업장별 배정 건수 ────────────────────────────────────────────────
const areaCounts = {{}};
BLOCKS.forEach(b => {{
  const a = b.assigned_area || '미배정';
  areaCounts[a] = (areaCounts[a] || 0) + 1;
}});
const areaLabels = Object.keys(areaCounts).sort();
const areaData   = areaLabels.map(a => areaCounts[a]);
new Chart(document.getElementById('areaChart'), {{
  type: 'bar',
  data: {{
    labels: areaLabels,
    datasets: [{{ label: '배정 건수', data: areaData,
      backgroundColor: '#0d6efd', borderRadius: 4
    }}]
  }},
  options: {{ plugins:{{legend:{{display:false}}}}, scales:{{y:{{beginAtZero:true}}}} }}
}});

// ── Chart: 조업도 월별 추이 ──────────────────────────────────────────────────
const months12 = {json.dumps([f'2026-{m:02d}' for m in range(1,13)])};
const areas    = Object.keys(STATS.utilization).sort();
const utilDs   = areas.map((area, i) => ({{
  label: area,
  data: months12.map(m => STATS.utilization[area][m] || 0),
  borderWidth: 1.5, pointRadius: 2, fill: false,
  borderColor: `hsl(${{(i * 360/areas.length).toFixed(0)}},65%,50%)`
}}));
new Chart(document.getElementById('utilChart'), {{
  type: 'line',
  data: {{ labels: months12, datasets: utilDs }},
  options: {{
    plugins: {{ legend: {{ position:'right', labels:{{ boxWidth:10, font:{{size:10}} }} }} }},
    scales: {{ y: {{ beginAtZero:true, title:{{display:true, text:'조업도(%)' }} }} }}
  }}
}});

// ── 작업장 필터 셀렉트 채우기 ────────────────────────────────────────────────
const fArea = document.getElementById('fArea');
[...new Set(BLOCKS.map(b => b.assigned_area).filter(Boolean))].sort()
  .forEach(a => {{ const o = new Option(a,a); fArea.add(o); }});

// ── 블록 목록 렌더링 ─────────────────────────────────────────────────────────
let filteredBlocks = [...BLOCKS];

function renderTable() {{
  const tbody = document.getElementById('blockBody');
  const ht    = document.getElementById('fHT').value;
  const stg   = document.getElementById('fStg').value;
  const rule  = document.getElementById('fRule').value;
  const area  = document.getElementById('fArea').value;
  const srch  = document.getElementById('fSearch').value.toLowerCase();

  filteredBlocks = BLOCKS.filter(b => {{
    if (ht   && b.h_t !== ht) return false;
    if (stg  && b.stg !== stg) return false;
    if (rule === 'NONE' && b.rule_assigned !== '') return false;
    if (rule && rule !== 'NONE' && b.rule_assigned !== rule) return false;
    if (area && b.assigned_area !== area) return false;
    if (srch && !b.blk.toLowerCase().includes(srch)) return false;
    return true;
  }});

  document.getElementById('rowCount').textContent = filteredBlocks.length.toLocaleString() + '건';

  const display = filteredBlocks.slice(0, 1000);
  tbody.innerHTML = display.map(b => `
    <tr>
      <td>${{b.pjt}}</td>
      <td><span class="badge ${{b.h_t==='H'?'bg-primary':'bg-success'}}">${{b.h_t}}</span></td>
      <td class="font-monospace">${{b.blk}}</td>
      <td><span class="badge ${{b.stg==='대'?'bg-danger':b.stg==='중'?'bg-warning text-dark':'bg-secondary'}}">${{b.stg}}</span></td>
      <td>${{b.m_stdt}}</td>
      <td>${{b.m_fndt}}</td>
      <td>${{b.m_area}}</td>
      <td class="text-end">${{b.load_mh?Number(b.load_mh).toFixed(1):''}}</td>
      <td><span class="badge bg-info text-dark">${{b.assigned_area||'<span style="color:#aaa">미배정</span>'}}</span></td>
      <td><span class="badge bg-secondary">${{b.rule_assigned||'-'}}</span></td>
    </tr>`).join('');
  if (filteredBlocks.length > 1000) {{
    tbody.innerHTML += `<tr><td colspan="10" class="text-center text-muted">
      ... 상위 1,000건 표시 중 (전체 ${{filteredBlocks.length.toLocaleString()}}건) ...</td></tr>`;
  }}
}}

['fHT','fStg','fRule','fArea'].forEach(id =>
  document.getElementById(id).addEventListener('change', renderTable));
document.getElementById('fSearch').addEventListener('input', renderTable);
renderTable();
</script>
</body>
</html>"""

    Path(REPORT_PATH).write_text(HTML, encoding='utf-8')
    print(f"🌐 HTML 리포트 생성 완료: {REPORT_PATH}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 메인
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

if __name__ == '__main__':
    agent_run_time = datetime.now()

    print("=" * 70)
    print("  블록 작업장 배정 Agent")
    print("=" * 70)

    print("\n▶ 데이터 로드 ...")
    df       = load_blk_master(BLK_MASTER_PATH)
    capa     = load_area_capa(AREA_CAPA_PATH)
    area_seq = load_area_seq(AREA_CAPA_PATH)
    print(f"  blk_master : {len(df):,}행")
    print(f"  area_capa  : {len(capa)}개 작업장 ({', '.join(sorted(capa.keys()))})")

    print("\n▶ 글로벌 부하 분산 계산 (load_mh → 월별/주별) ...")
    load_dists = [calc_load_distribution(row) for _, row in df.iterrows()]

    # 누적 부하 추적 테이블 (월별 + 주별)
    all_areas = list(capa.keys()) + [AREA_Y9]
    cl:  dict = {a: defaultdict(float) for a in all_areas}
    cwl: dict = {a: defaultdict(float) for a in all_areas}

    print("\n▶ 순차 배정 규칙 실행 ...")
    r0_report_transfer_capacity(df, load_dists, capa)
    tgt = r1_calc_target_loads(df, load_dists, capa)
    r2_assign_y9(df, load_dists, cl, cwl)
    r3_assign_t_mid_large(df, load_dists, cl, cwl)
    r4_assign_t_small_dongil(df, load_dists, cl, cwl)
    r5_report_h_small(df, load_dists, cl, capa)
    r6_assign_h_small(df, load_dists, cl, cwl, tgt)
    r7_assign_h_mid_jig_f(df, load_dists, cl, cwl, tgt)
    r8_assign_h_mid_jig_dl(df, load_dists, cl, cwl, tgt)
    r9_assign_specialized_family(df, load_dists, cl, cwl, tgt)
    r10_assign_h_large_jig_dl(df, load_dists, cl, cwl, tgt)
    r11_assign_h_remaining(df, load_dists, cl, cwl, tgt)
    r12_assign_h_dir(df, load_dists, cl, cwl)
    r13_assign_t_small_dist(df, load_dists, cl, cwl, tgt, area_seq)

    print_utilization_summary(df, cl, capa)
    stats = save_stats(df, cl, capa, tgt)
    save_result(df, RESULT_PATH)
    generate_html_report(df, stats, agent_run_time)
