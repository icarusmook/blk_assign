"""
app.py — 블록 작업장 배정 + Streamlit 시각화 통합
(blk_assign_agent.py 배정 로직 + app.py UI, HTML 리포트 생성 제외)
"""

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Imports
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

import io
import json
import contextlib
from datetime import date, timedelta
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import plotly.express as px
import streamlit as st
import streamlit.components.v1 as components

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 경로 상수
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

BASE_DIR        = Path(__file__).parent
BLK_MASTER_PATH = BASE_DIR / "blk_master.xlsx"
AREA_CAPA_PATH  = BASE_DIR / "area_capa.xlsx"
RESULT_PATH     = BASE_DIR / "blk_assign_result.xlsx"
STATS_PATH      = BASE_DIR / "blk_assign_stats.json"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 배정 로직 상수
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

KR_HOLIDAYS = {
    date(2026,  1,  1),
    date(2026,  2, 16), date(2026,  2, 17), date(2026,  2, 18),
    date(2026,  3,  1),
    date(2026,  5,  5), date(2026,  5, 26),
    date(2026,  6,  6),
    date(2026,  8, 15),
    date(2026,  9, 25), date(2026,  9, 26), date(2026,  9, 27), date(2026,  9, 28),
    date(2026, 10,  3), date(2026, 10,  9),
    date(2026, 12, 25),
    date(2027,  1,  1), date(2027,  1, 27), date(2027,  1, 28), date(2027,  1, 29),
    date(2027,  3,  1),
}

AREA_Y9       = 'AREA15'
AREAS_SMALL   = ['AREA3', 'AREA4', 'AREA5']
AREAS_MID_DL  = ['AREA1', 'AREA2', 'AREA6']
AREAS_REMAIN  = ['AREA7', 'AREA8', 'AREA9', 'AREA10', 'AREA12', 'AREA13', 'AREA14']
PS_AREA_PAIRS = [('AREA7', 'AREA8'), ('AREA9', 'AREA10'), ('AREA13', 'AREA14')]
AREA_PRIOR_COLS = ['area_prior_1', 'area_prior_2', 'area_prior_3', 'area_prior_4', 'area_prior_5']

RULE_LABELS = {
    'R2':  'H/Y9 → AREA15',
    'R3':  'T-중/대 → m_area',
    'R4':  'T-소/동일 → prt_area',
    'R6':  'H-소 → AREA3~5 순환',
    'R7':  'H-중-jig_F → AREA2',
    'R8':  'H-중-jig_D/L → AREA1→2→6',
    'R9':  'H-전문화 Family',
    'R10': 'H-대-jig_D/L 그룹 → AREA12',
    'R11': 'H-대/중 잔여 → AREA7~14',
    'R12': 'H-dir → prt_area(override)',
    'R13': 'T-소 분산배정 (area_seq 순)',
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 유틸리티
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def is_workday(d: date) -> bool:
    return d.weekday() < 5 and d not in KR_HOLIDAYS

def get_workdays_between(start: date, end: date) -> list:
    days, cur = [], start
    while cur <= end:
        if is_workday(cur):
            days.append(cur)
        cur += timedelta(days=1)
    return days

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 데이터 로드
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def load_blk_master(path) -> pd.DataFrame:
    df = pd.read_excel(path)
    df['m_stdt'] = pd.to_datetime(df['m_stdt'])
    df['m_fndt'] = pd.to_datetime(df['m_fndt'])
    df['assigned_area'] = None
    df['rule_assigned']  = ''
    return df

def load_area_capa(path) -> dict:
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    capa = {}
    for row in rows[2:]:
        area = row[0]
        if not area:
            continue
        capa[area] = {m: (row[m] or 0) for m in range(1, 13)}
    return capa

def load_area_seq(path) -> list:
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    areas = []
    for row in rows[2:]:
        if row[0]:
            seq = row[13] if (len(row) > 13 and row[13] is not None) else 999
            areas.append((int(seq), row[0]))
    return [name for _, name in sorted(areas)]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 글로벌: 부하 분산 계산
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def calc_load_distribution(row: pd.Series) -> dict:
    load_mh = row['load_mh']
    m_stdt  = row['m_stdt']
    m_fndt  = row['m_fndt']
    if pd.isna(load_mh) or pd.isna(m_stdt) or pd.isna(m_fndt):
        return {'monthly': {}, 'weekly': {}, 'daily_mh': 0.0}
    workdays = get_workdays_between(m_stdt.date(), m_fndt.date())
    if not workdays:
        return {'monthly': {}, 'weekly': {}, 'daily_mh': 0.0}
    daily_mh = load_mh / len(workdays)
    monthly, weekly = defaultdict(float), defaultdict(float)
    for d in workdays:
        monthly[(d.year, d.month)] += daily_mh
        iso = d.isocalendar()
        weekly[(iso.year, iso.week)] += daily_mh
    return {'monthly': dict(monthly), 'weekly': dict(weekly), 'daily_mh': daily_mh}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 글로벌 헬퍼: 부하 추적 & 배정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _init_area(area, cl, cwl):
    if area not in cl:
        cl[area]  = defaultdict(float)
        cwl[area] = defaultdict(float)

def add_load(area, dist, cl, cwl):
    _init_area(area, cl, cwl)
    for k, v in dist['monthly'].items():
        cl[area][k] += v
    for k, v in dist['weekly'].items():
        cwl[area][k] += v

def remove_load(area, dist, cl, cwl):
    for k, v in dist['monthly'].items():
        cl[area][k] -= v
    for k, v in dist['weekly'].items():
        cwl[area][k] -= v

def can_assign(area, dist, cl, tgt) -> bool:
    if area not in tgt:
        return True
    for (yr, mo), blk_load in dist['monthly'].items():
        if cl.get(area, {}).get((yr, mo), 0.0) + blk_load > tgt[area].get((yr, mo), 0):
            return False
    return True

def can_assign_group(area, indices, load_dists, cl, tgt) -> bool:
    if area not in tgt:
        return True
    grp_load: dict = defaultdict(float)
    for idx in indices:
        for k, v in load_dists[idx]['monthly'].items():
            grp_load[k] += v
    for (yr, mo), g_load in grp_load.items():
        if cl.get(area, {}).get((yr, mo), 0.0) + g_load > tgt[area].get((yr, mo), 0):
            return False
    return True

def assign_block(df, idx, area, rule, load_dists, cl, cwl):
    df.at[idx, 'assigned_area'] = area
    df.at[idx, 'rule_assigned']  = rule
    add_load(area, load_dists[idx], cl, cwl)

def assign_roundrobin(candidates, areas, df, rule, load_dists, cl, cwl, tgt) -> int:
    area_idx = 0
    cnt = 0
    skip_streak = 0
    max_skip = len(areas)
    for idx in candidates:
        assigned = False
        for offset in range(len(areas)):
            area = areas[(area_idx + offset) % len(areas)]
            if can_assign(area, load_dists[idx], cl, tgt):
                assign_block(df, idx, area, rule, load_dists, cl, cwl)
                area_idx = (area_idx + offset + 1) % len(areas)
                cnt += 1
                skip_streak = 0
                assigned = True
                break
        if not assigned:
            skip_streak += 1
            if skip_streak >= max_skip * 3:
                break
    return cnt

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 배정 규칙 R0 ~ R13
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def r0_report_transfer_capacity(df, load_dists, capa):
    print("\n[R0] 이관물량(H) 처리가능여부 리포팅")
    year = 2026
    t_load: dict = defaultdict(float)
    h_load: dict = defaultdict(float)
    for i, row in df.iterrows():
        for (yr, mo), load in load_dists[i]['monthly'].items():
            if yr != year:
                continue
            if row['h_t'] == 'T':
                t_load[mo] += load
            else:
                h_load[mo] += load
    total_capa = {mo: sum(c.get(mo, 0) for a, c in capa.items() if a != AREA_Y9)
                  for mo in range(1, 13)}
    alert = False
    for mo in range(1, 13):
        t, h = t_load.get(mo, 0), h_load.get(mo, 0)
        total = t + h * 1.2
        cap   = total_capa.get(mo, 0)
        if total > cap:
            if not alert:
                print("  ⚠️ 처리 초과 월:")
                print(f"  {'월':^7} {'T부하':>10} {'H부하×1.2':>12} {'합계':>10} {'전체능력':>10} {'초과':>10}")
                alert = True
            print(f"  {year}-{mo:02d}  {t:>10,.0f} {h*1.2:>12,.0f} {total:>10,.0f} {cap:>10,.0f} {total-cap:>10,.0f}")
    if not alert:
        print("  ✅ 전체 월 처리 가능 (초과 없음)")


def r1_calc_target_loads(df, load_dists, capa) -> dict:
    year = 2026
    total_load: dict = defaultdict(float)
    for dist in load_dists:
        for (yr, mo), load in dist['monthly'].items():
            if yr == year:
                total_load[(yr, mo)] += load
    total_capa = {(year, mo): sum(c.get(mo, 0) for a, c in capa.items() if a != AREA_Y9)
                  for mo in range(1, 13)}
    util = {k: (total_load[k] / total_capa[k] if total_capa[k] > 0 else 0)
            for k in total_capa}
    tgt: dict = {}
    for area, month_caps in capa.items():
        if area == AREA_Y9:
            continue
        tgt[area] = {(year, mo): month_caps.get(mo, 0) * util.get((year, mo), 0)
                     for mo in range(1, 13)}
    print(f"\n[R1] 목표물량 산출 완료 (목표조업도 {util.get((year,1),0)*100:.1f}%~)")
    return tgt


def r2_assign_y9(df, load_dists, cl, cwl):
    mask = (df['h_t'] == 'H') & df['blk'].str.startswith('Y9') & df['assigned_area'].isna()
    cnt = 0
    for idx in df[mask].index:
        assign_block(df, idx, AREA_Y9, 'R2', load_dists, cl, cwl)
        cnt += 1
    print(f"  [R2] H/Y9 → AREA15: {cnt}건")


def r3_assign_t_mid_large(df, load_dists, cl, cwl):
    mask = (df['h_t'] == 'T') & df['stg'].isin(['중', '대']) & df['assigned_area'].isna()
    cnt = 0
    for idx in df[mask].index:
        area = df.at[idx, 'm_area']
        if pd.notna(area) and area:
            assign_block(df, idx, area, 'R3', load_dists, cl, cwl)
            cnt += 1
    print(f"  [R3] T-중/대 → m_area: {cnt}건")


def r4_assign_t_small_dongil(df, load_dists, cl, cwl):
    mask = (
        (df['h_t'] == 'T') & (df['stg'] == '소') &
        (df['area_prior_1'] == '동일') & df['assigned_area'].isna()
    )
    cnt = 0
    for idx in df[mask].index:
        area = df.at[idx, 'prt_area']
        if pd.notna(area) and area:
            assign_block(df, idx, area, 'R4', load_dists, cl, cwl)
            cnt += 1
    print(f"  [R4] T-소/동일 → prt_area: {cnt}건")


def r5_report_h_small(df, load_dists, cl, capa):
    print("\n[R5] H-소 물량 AREA3~5 처리가능여부 리포팅")
    year = 2026
    h_so_load: dict = defaultdict(float)
    mask = (df['h_t'] == 'H') & (df['stg'] == '소') & df['assigned_area'].isna()
    for i in df[mask].index:
        for (yr, mo), load in load_dists[i]['monthly'].items():
            if yr == year:
                h_so_load[mo] += load
    area35_remain: dict = defaultdict(float)
    for area in AREAS_SMALL:
        for mo in range(1, 13):
            cap = capa.get(area, {}).get(mo, 0)
            cur = cl.get(area, {}).get((year, mo), 0.0)
            area35_remain[mo] += max(0.0, cap - cur)
    alert = False
    for mo in range(1, 13):
        h   = h_so_load.get(mo, 0)
        rem = area35_remain.get(mo, 0)
        if h > rem:
            if not alert:
                print("  ⚠️ 초과 월:")
                print(f"  {'월':^7} {'H-소부하':>12} {'AREA3~5잔여':>14} {'초과':>10}")
                alert = True
            print(f"  {year}-{mo:02d}  {h:>12,.0f} {rem:>14,.0f} {h-rem:>10,.0f}")
    if not alert:
        print("  ✅ AREA3~5에서 H-소 전량 처리 가능")


def r6_assign_h_small(df, load_dists, cl, cwl, tgt):
    mask  = (df['h_t'] == 'H') & (df['stg'] == '소') & df['assigned_area'].isna()
    cands = df[mask].sort_values('m_stdt').index.tolist()
    cnt   = assign_roundrobin(cands, AREAS_SMALL, df, 'R6', load_dists, cl, cwl, tgt)
    print(f"  [R6] H-소 → AREA3~5 순환: {cnt}건")


def r7_assign_h_mid_jig_f(df, load_dists, cl, cwl, tgt):
    area = 'AREA2'
    SPECIAL_BLK = {'D114', 'D174', 'D204'}
    mask_p1 = (
        (df['h_t'] == 'H') & (df['ship_kind'] == 'A') & (df['stg'] == '중') &
        df['blk'].str[:4].isin(SPECIAL_BLK) & df['assigned_area'].isna()
    )
    cnt_p1 = 0
    for idx in df[mask_p1].index:
        assign_block(df, idx, area, 'R7', load_dists, cl, cwl)
        cnt_p1 += 1
    mask_p2 = (
        (df['h_t'] == 'H') & (df['stg'] == '중') &
        (df['jig'] == 'F') & df['assigned_area'].isna()
    )
    cnt_p2 = 0
    skip_streak = 0
    for idx in df[mask_p2].sort_values('m_stdt').index:
        if can_assign(area, load_dists[idx], cl, tgt):
            assign_block(df, idx, area, 'R7', load_dists, cl, cwl)
            cnt_p2 += 1
            skip_streak = 0
        else:
            skip_streak += 1
            if skip_streak >= 50:
                break
    print(f"  [R7] H-중-jig_F → AREA2: Part1={cnt_p1}건 + Part2={cnt_p2}건")


def r8_assign_h_mid_jig_dl(df, load_dists, cl, cwl, tgt):
    mask = (
        (df['h_t'] == 'H') & (df['stg'] == '중') &
        ~df['blk'].str.startswith('H') &
        ~df['blk'].str[:3].isin(['E11', 'E51']) &
        df['jig'].isin(['D', 'L', 'W']) & df['assigned_area'].isna()
    )
    cands = df[mask].sort_values('m_stdt').index.tolist()
    cnt   = assign_roundrobin(cands, AREAS_MID_DL, df, 'R8', load_dists, cl, cwl, tgt)
    print(f"  [R8] H-중-jig_D/L/W → AREA1→2→6 순환: {cnt}건")


def r9_assign_specialized_family(df, load_dists, cl, cwl, tgt):
    def _assign(blk_filter, suffix, area, label):
        base = (df['h_t'] == 'H') & df['stg'].isin(['대', '중']) & df['assigned_area'].isna()
        sub  = df[base & blk_filter & (df['blk'].str[-1] == suffix)].copy()
        sub['_grp'] = sub['pjt'].astype(str) + '|' + sub['blk'].str[:3]
        grp_order = sub.groupby('_grp')['m_stdt'].min().sort_values().index.tolist()
        cnt = 0
        skip_streak = 0
        for grp in grp_order:
            g_idx = sub[sub['_grp'] == grp].index.tolist()
            for idx in df.loc[g_idx].sort_values('m_stdt').index:
                if df.at[idx, 'assigned_area'] is not None:
                    continue
                if can_assign(area, load_dists[idx], cl, tgt):
                    assign_block(df, idx, area, 'R9', load_dists, cl, cwl)
                    cnt += 1
                    skip_streak = 0
                else:
                    skip_streak += 1
                    if skip_streak >= 50:
                        print(f"  [R9-{label}] → {area}: {cnt}건")
                        return
        print(f"  [R9-{label}] → {area}: {cnt}건")

    _assign(df['blk'].str.startswith('H'),             'P', 'AREA7',  '1-1')
    _assign(df['blk'].str.startswith('H'),             'S', 'AREA8',  '1-2')
    _assign(df['blk'].str[:3].isin(['E11', 'F51']),    'P', 'AREA9',  '2-1')
    _assign(df['blk'].str[:3].isin(['E11', 'F51']),    'S', 'AREA10', '2-2')


def r10_assign_h_large_jig_dl(df, load_dists, cl, cwl, tgt):
    area = 'AREA12'
    seed_mask = (
        (df['h_t'] == 'H') & (df['stg'] == '대') &
        df['jig'].isin(['D', 'L']) & df['assigned_area'].isna()
    )
    grp_keys = set(
        df[seed_mask].apply(lambda r: (r['pjt'], r['blk'][:3], r['blk'][-1]), axis=1)
    )
    def in_grp(r):
        return (r['pjt'], r['blk'][:3], r['blk'][-1]) in grp_keys
    all_h_unassigned = (df['h_t'] == 'H') & df['assigned_area'].isna()
    cands = (df[all_h_unassigned][df[all_h_unassigned].apply(in_grp, axis=1)]
             .sort_values('m_stdt').index.tolist())
    cnt = 0
    skip_streak = 0
    for idx in cands:
        if can_assign(area, load_dists[idx], cl, tgt):
            assign_block(df, idx, area, 'R10', load_dists, cl, cwl)
            cnt += 1
            skip_streak = 0
        else:
            skip_streak += 1
            if skip_streak >= 50:
                break
    print(f"  [R10] H-대-jig_D/L 그룹 → AREA12: {cnt}건")


def r11_assign_h_remaining(df, load_dists, cl, cwl, tgt):
    mask = (df['h_t'] == 'H') & df['stg'].isin(['대', '중']) & df['assigned_area'].isna()
    sub  = df[mask].copy()
    sub['_grp_key'] = sub['pjt'].astype(str) + '|' + sub['blk'].str[:3] + '|' + sub['blk'].str[-1]
    sub['_ps_key']  = sub['pjt'].astype(str) + '|' + sub['blk'].str[:3]
    sfx_by_ps = sub.groupby('_ps_key')['_grp_key'].apply(lambda ks: {k.split('|')[2] for k in ks})
    ps_pairs  = {k for k, s in sfx_by_ps.items() if 'P' in s and 'S' in s}
    grp_min_stdt = sub.groupby('_grp_key')['m_stdt'].min().sort_values()
    processed: set = set()
    cnt = 0

    def assign_individually(idx_list):
        nonlocal cnt
        for i in sorted(idx_list, key=lambda x: df.at[x, 'm_stdt']):
            if df.at[i, 'assigned_area'] is not None:
                continue
            for a in AREAS_REMAIN:
                if can_assign(a, load_dists[i], cl, tgt):
                    assign_block(df, i, a, 'R11', load_dists, cl, cwl)
                    cnt += 1
                    break

    for grp_key in grp_min_stdt.index:
        if grp_key in processed:
            continue
        parts = grp_key.split('|')
        pjt_v, blk3, sfx = int(parts[0]), parts[1], parts[2]
        ps_key    = f"{pjt_v}|{blk3}"
        unassigned = [i for i in sub[sub['_grp_key'] == grp_key].index
                      if df.at[i, 'assigned_area'] is None]
        if not unassigned:
            processed.add(grp_key)
            continue
        if ps_key in ps_pairs and sfx in ('P', 'S'):
            partner_sfx = 'S' if sfx == 'P' else 'P'
            partner_key = f"{pjt_v}|{blk3}|{partner_sfx}"
            processed.add(grp_key)
            processed.add(partner_key)
            p_key = grp_key     if sfx == 'P' else partner_key
            s_key = partner_key if sfx == 'P' else grp_key
            p_idx = [i for i in sub[sub['_grp_key'] == p_key].index
                     if df.at[i, 'assigned_area'] is None]
            s_idx = [i for i in sub[sub['_grp_key'] == s_key].index
                     if df.at[i, 'assigned_area'] is None]
            pair_done = False
            for (p_area, s_area) in PS_AREA_PAIRS:
                if (can_assign_group(p_area, p_idx, load_dists, cl, tgt) and
                        can_assign_group(s_area, s_idx, load_dists, cl, tgt)):
                    for i in p_idx:
                        assign_block(df, i, p_area, 'R11', load_dists, cl, cwl)
                        cnt += 1
                    for i in s_idx:
                        assign_block(df, i, s_area, 'R11', load_dists, cl, cwl)
                        cnt += 1
                    pair_done = True
                    break
            if not pair_done:
                assign_individually(p_idx + s_idx)
        else:
            processed.add(grp_key)
            grp_done = False
            for a in AREAS_REMAIN:
                if can_assign_group(a, unassigned, load_dists, cl, tgt):
                    for i in unassigned:
                        assign_block(df, i, a, 'R11', load_dists, cl, cwl)
                        cnt += 1
                    grp_done = True
                    break
            if not grp_done:
                assign_individually(unassigned)
    print(f"  [R11] H-대/중 잔여 → AREA7~14: {cnt}건")


def r12_assign_h_dir(df, load_dists, cl, cwl):
    mask = (df['h_t'] == 'H') & df['blk'].str[3].isin(['9', 'H'])
    cnt = 0
    for idx in df[mask].index:
        area = df.at[idx, 'prt_area']
        if not (pd.notna(area) and area):
            continue
        old = df.at[idx, 'assigned_area']
        if old and old != area:
            remove_load(old, load_dists[idx], cl, cwl)
        assign_block(df, idx, area, 'R12', load_dists, cl, cwl)
        cnt += 1
    print(f"  [R12] H-dir(blk[3]=9/H) → prt_area override: {cnt}건")


def r13_assign_t_small_dist(df, load_dists, cl, cwl, tgt, area_seq):
    total_cnt = 0
    for area in area_seq:
        if area not in tgt:
            continue
        area_cnt = 0
        for prior_col in AREA_PRIOR_COLS:
            sub_mask = (
                (df['h_t'] == 'T') & (df['stg'] == '소') &
                df['assigned_area'].isna() & (df[prior_col] == area)
            )
            sub = df[sub_mask].copy()
            if sub.empty:
                continue
            sub['_grp'] = (sub['pjt'].astype(str) + '|' +
                           sub['blk'].astype(str) + '|' +
                           sub['type'].astype(str))
            grp_info = (sub.groupby('_grp')
                        .agg(ap=('assign_prior', 'min'), ms=('m_stdt', 'min'))
                        .sort_values(['ap', 'ms']))
            grp_order = grp_info.index.tolist()
            area_skip = 0
            for grp_key in grp_order:
                grp_idx = sub[sub['_grp'] == grp_key].sort_values('m_stdt').index.tolist()
                for idx in grp_idx:
                    if df.at[idx, 'assigned_area'] is not None:
                        continue
                    if can_assign(area, load_dists[idx], cl, tgt):
                        assign_block(df, idx, area, 'R13', load_dists, cl, cwl)
                        area_cnt += 1
                        total_cnt += 1
                        area_skip = 0
                    else:
                        area_skip += 1
                if area_skip >= 50:
                    break
        print(f"    [R13] {area}: {area_cnt}건")
    print(f"  [R13] T-소 분산배정 합계: {total_cnt}건")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 통계 생성 & 파일 저장
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_stats(df, cl, capa, tgt) -> dict:
    year = 2026
    steps = []
    for rule in ['R2','R3','R4','R6','R7','R8','R9','R10','R11','R12','R13']:
        sub = df[df['rule_assigned'] == rule]
        area_dist = sub['assigned_area'].value_counts().to_dict()
        steps.append({
            'rule':  rule,
            'label': RULE_LABELS.get(rule, rule),
            'count': int(len(sub)),
            'areas': {k: int(v) for k, v in area_dist.items()},
        })
    utilization = {}
    for area in sorted(a for a in capa if a != AREA_Y9):
        utilization[area] = {}
        for mo in range(1, 13):
            cap  = capa[area].get(mo, 0)
            load = cl.get(area, {}).get((year, mo), 0.0)
            utilization[area][f'{year}-{mo:02d}'] = round(load / cap * 100, 1) if cap > 0 else 0.0
    target_util = {}
    for area in sorted(a for a in capa if a != AREA_Y9):
        target_util[area] = {}
        for mo in range(1, 13):
            cap = capa[area].get(mo, 0)
            t   = tgt.get(area, {}).get((year, mo), 0)
            target_util[area][f'{year}-{mo:02d}'] = round(t / cap * 100, 1) if cap > 0 else 0.0
    return {
        'total':       int(len(df)),
        'assigned':    int(df['assigned_area'].notna().sum()),
        'unassigned':  int(df['assigned_area'].isna().sum()),
        'steps':       steps,
        'utilization': utilization,
        'target_util': target_util,
    }

def save_files(df, stats):
    with open(STATS_PATH, 'w', encoding='utf-8') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    out = df.copy()
    out['m_stdt'] = out['m_stdt'].dt.date
    out['m_fndt'] = out['m_fndt'].dt.date
    out.to_excel(RESULT_PATH, index=False)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 메인 배정 실행 (캐시)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@st.cache_data(show_spinner=False)
def run_assignment(blk_mtime: float, capa_mtime: float):
    """입력 파일 수정시간이 바뀔 때만 재실행. stdout을 캡처해서 로그로 반환.

    blk_mtime, capa_mtime 은 @st.cache_data 의 캐시 키로만 사용되며,
    함수 본문에서 직접 참조하지 않는 것이 정상입니다.
    """
    del blk_mtime, capa_mtime  # cache keys only — consumed by @st.cache_data
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        print("▶ 데이터 로드 ...")
        df       = load_blk_master(BLK_MASTER_PATH)
        capa     = load_area_capa(AREA_CAPA_PATH)
        area_seq = load_area_seq(AREA_CAPA_PATH)
        print(f"  blk_master: {len(df):,}행  |  area_capa: {len(capa)}개 작업장")

        print("\n▶ 부하 분산 계산 ...")
        load_dists = [calc_load_distribution(row) for _, row in df.iterrows()]

        all_areas = list(capa.keys()) + [AREA_Y9]
        cl:  dict = {a: defaultdict(float) for a in all_areas}
        cwl: dict = {a: defaultdict(float) for a in all_areas}

        print("\n▶ 순차 배정 규칙 실행 ...")
        r0_report_transfer_capacity(df, load_dists, capa)
        tgt = r1_calc_target_loads(df, load_dists, capa)
        r2_assign_y9(df, load_dists, cl, cwl)
        r3_assign_t_mid_large(df, load_dists, cl, cwl)
        r4_assign_t_small_dongil(df, load_dists, cl, cwl)
        r5_report_h_small(df, load_dists, cl, capa)
        r6_assign_h_small(df, load_dists, cl, cwl, tgt)
        r7_assign_h_mid_jig_f(df, load_dists, cl, cwl, tgt)
        r8_assign_h_mid_jig_dl(df, load_dists, cl, cwl, tgt)
        r9_assign_specialized_family(df, load_dists, cl, cwl, tgt)
        r10_assign_h_large_jig_dl(df, load_dists, cl, cwl, tgt)
        r11_assign_h_remaining(df, load_dists, cl, cwl, tgt)
        r12_assign_h_dir(df, load_dists, cl, cwl)
        r13_assign_t_small_dist(df, load_dists, cl, cwl, tgt, area_seq)

        stats = build_stats(df, cl, capa, tgt)
        save_files(df, stats)
        asgn = stats['assigned']
        print(f"\n✅ 배정 완료: {asgn:,}/{len(df):,}건 ({asgn/len(df)*100:.1f}%)")

    return df, stats, buf.getvalue()


def get_file_mtime(path: Path) -> float:
    return path.stat().st_mtime if path.exists() else 0.0

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# UI 상수
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

RULE_META = {
    "R0": {
        "label": "이관물량 처리가능여부 리포팅",
        "badge": "warning",
        "desc": "T부하 + H부하×1.2 vs 전체능력 비교 → 초과 월 경고 출력 (배정 없음)",
        "detail": "월별 부하 총합(T선 부하 + H선 부하×1.2)이 전체 작업장 능력을 초과하는 경우 경고를 출력합니다. 실제 블록 배정은 수행하지 않는 리포팅 전용 단계입니다.",
    },
    "R1": {
        "label": "작업장별 월별 배정 목표물량 산출",
        "badge": "success",
        "desc": "목표조업도 = Σ전체부하 ÷ Σ전체능력 → 배정목표 = 목표조업도 × 각 작업장 능력",
        "detail": "전체 부하를 전체 능력으로 나눈 목표조업도를 산출하고, 각 작업장·월별 배정 목표물량(MH)을 계산합니다. 이후 모든 규칙의 배정 상한선 기준이 됩니다.",
    },
    "R2": {
        "label": "H/Y9 블록 → AREA15",
        "badge": "primary",
        "desc": "h_t=H이고 blk가 Y9로 시작하는 블록 → AREA15 무조건 배정 (목표물량 무관)",
        "detail": "블록 번호가 Y9로 시작하는 H행 블록은 목표물량과 무관하게 AREA15로 우선 배정됩니다.",
    },
    "R3": {
        "label": "T-중/대 → m_area",
        "badge": "primary",
        "desc": "h_t=T, stg=중 또는 대 → m_area 컬럼 값으로 직접 배정",
        "detail": "T선 중형·대형 블록은 마스터 데이터의 m_area(지정 작업장)에 직접 배정합니다.",
    },
    "R4": {
        "label": "T-소/동일 → prt_area",
        "badge": "primary",
        "desc": "h_t=T, stg=소, area_prior_1=동일 → prt_area 컬럼 값으로 배정",
        "detail": "T선 소형 블록 중 area_prior_1이 '동일'로 지정된 경우 prt_area(부품 작업장)에 배정합니다.",
    },
    "R5": {
        "label": "H-소 처리가능여부 리포팅",
        "badge": "warning",
        "desc": "H-소 부하 vs AREA3~5 잔여능력 비교 → 초과 월 경고 출력 (배정 없음)",
        "detail": "H선 소형 블록의 월별 부하가 AREA3~5의 잔여 능력을 초과하는 경우 경고를 출력합니다. 리포팅 전용이며 실제 배정은 수행하지 않습니다.",
    },
    "R6": {
        "label": "H-소 → AREA3~5 순환",
        "badge": "primary",
        "desc": "h_t=H, stg=소 → m_stdt 오름차순 정렬 후 AREA3→AREA4→AREA5 순환. 목표 도달 시 해당 area 중단",
        "detail": "H선 소형 블록을 착수일 순으로 정렬하여 AREA3, AREA4, AREA5에 라운드로빈 방식으로 배정합니다.",
    },
    "R7": {
        "label": "H-중-jig_F → AREA2",
        "badge": "primary",
        "desc": "Part1: 특정 블록(D114·D174·D204) 무조건 AREA2 / Part2: 나머지 jig=F 목표물량까지 AREA2",
        "detail": "H선 중형 jig=F 블록을 AREA2에 배정합니다. Part1은 특정 블록 번호를 무조건 배정하고, Part2는 목표물량 한도 내에서 나머지 jig_F 블록을 배정합니다.",
    },
    "R8": {
        "label": "H-중-jig_D/L → AREA1→2→6 순환",
        "badge": "primary",
        "desc": "h_t=H, stg=중, jig=D 또는 L (단 블록 번호 H·E11·E51 시작 제외) → AREA1→AREA2→AREA6 순환",
        "detail": "H선 중형 D/L jig 블록을 AREA1, AREA2, AREA6 순환 배정합니다. 단, 블록 번호가 H, E11, E51로 시작하는 블록은 제외합니다.",
    },
    "R9": {
        "label": "H-전문화 Family 배정",
        "badge": "primary",
        "desc": "pjt+blk앞3자리로 그룹 / P말미→AREA7(H) 또는 AREA9(E11·F51) / S말미→AREA8(H) 또는 AREA10(E11·F51)",
        "detail": "H선 대·중형 전문화 패밀리 블록을 블록 말미 알파벳으로 분류합니다. P 말미이면 AREA7 또는 AREA9, S 말미이면 AREA8 또는 AREA10에 배정합니다.",
    },
    "R10": {
        "label": "H-대-jig_D/L 그룹 → AREA12",
        "badge": "primary",
        "desc": "h_t=H, stg=대, jig=D 또는 L → 씨드블록과 동일 (pjt+blk앞3자리+말미) 그룹으로 AREA12 배정",
        "detail": "H선 대형 D/L jig 블록을 그룹 단위로 AREA12에 배정합니다. 목표물량 도달 시 중단합니다.",
    },
    "R11": {
        "label": "H-대/중 잔여 → AREA7~14",
        "badge": "primary",
        "desc": "PS페어 그룹: (AREA7+8), (AREA9+10), (AREA13+14) 우선 / 단독: AREA7→14 first-fit / 초과 시 개별 배정",
        "detail": "H선 대·중형 잔여 블록을 PS 페어 작업장에 우선 배정합니다. 페어 배정이 불가능하면 AREA7~14를 순차 탐색(first-fit)하며, 모두 초과 시 개별 배정합니다.",
    },
    "R12": {
        "label": "H-dir → prt_area override",
        "badge": "danger",
        "desc": "h_t=H, blk 4번째 자리가 9 또는 H → prt_area로 강제 override 배정",
        "detail": "블록 번호의 4번째 자리가 9 또는 H인 H선 블록은 이전 규칙으로 배정된 작업장을 무시하고 prt_area로 강제 변경합니다.",
    },
    "R13": {
        "label": "T-소 분산배정",
        "badge": "primary",
        "desc": "area_seq 순으로 분산 / 그룹: pjt+blk+type / 정렬: assign_prior→m_stdt / area_prior_1~5 단계적 확장",
        "detail": "T선 소형 미배정 잔여 블록을 area_seq 순서에 따라 분산 배정합니다. area_prior_1부터 5까지 단계적으로 확장하며 목표물량 도달 시 다음 area로 넘어갑니다.",
    },
}

BADGE_COLOR = {"primary": "#0d6efd", "success": "#198754", "warning": "#856404", "danger": "#dc3545"}
BADGE_BG    = {"primary": "#cfe2ff", "success": "#d1e7dd", "warning": "#fff3cd", "danger": "#f8d7da"}

def area_key(x):
    return int(x.replace("AREA", "")) if isinstance(x, str) and x.startswith("AREA") else 99

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Streamlit 앱
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

st.set_page_config(page_title="블록 배정 대시보드", page_icon="🏗️", layout="wide")

# ── 사이드바 ──────────────────────────────────────────────────────────────────

st.sidebar.header("🏗️ 블록 작업장 배정")

if st.sidebar.button("🔄 배정 재실행", use_container_width=True):
    run_assignment.clear()
    st.rerun()

st.sidebar.divider()
st.sidebar.header("필터 (조업도·상세 데이터 탭 적용)")

# ── 배정 실행 ──────────────────────────────────────────────────────────────────

blk_mtime  = get_file_mtime(BLK_MASTER_PATH)
capa_mtime = get_file_mtime(AREA_CAPA_PATH)

with st.spinner("배정 로직 실행 중..."):
    df_all, stats, run_log = run_assignment(blk_mtime, capa_mtime)

# ── 사이드바 필터 ─────────────────────────────────────────────────────────────

pjt_opts  = sorted(df_all["pjt"].dropna().unique().tolist())
sel_pjt   = st.sidebar.multiselect("프로젝트(pjt)", pjt_opts, default=pjt_opts)

ht_opts   = sorted(df_all["h_t"].dropna().unique().tolist())
sel_ht    = st.sidebar.multiselect("선종 구분(H/T)", ht_opts, default=ht_opts)

area_opts = sorted(df_all["assigned_area"].dropna().unique().tolist(), key=area_key)
sel_area  = st.sidebar.multiselect("배정 작업장", area_opts, default=area_opts)

stg_opts  = sorted(df_all["stg"].dropna().unique().tolist())
sel_stg   = st.sidebar.multiselect("크기(stg)", stg_opts, default=stg_opts)

st.sidebar.divider()
with st.sidebar.expander("실행 로그"):
    st.text(run_log)

df_base     = df_all[df_all["pjt"].isin(sel_pjt) & df_all["h_t"].isin(sel_ht) & df_all["stg"].isin(sel_stg)]
df_filtered = df_base[df_base["assigned_area"].isin(sel_area) | df_base["assigned_area"].isna()]

# ── 헤더 + KPI ────────────────────────────────────────────────────────────────

st.title("🏗️ 블록 작업장 배정 대시보드")

total     = stats["total"]
assigned  = stats["assigned"]
unassigned = stats["unassigned"]
rate      = round(assigned / total * 100, 1) if total else 0

c1, c2, c3, c4 = st.columns(4)
c1.metric("전체 블록", f"{total:,}")
c2.metric("배정 완료", f"{assigned:,}", f"{rate}%")
c3.metric("미배정", f"{unassigned:,}", f"-{round(unassigned/total*100,1)}%", delta_color="inverse")
c4.metric("배정 규칙 수", "11", "R2 ~ R13")

st.divider()

# ── 탭 ───────────────────────────────────────────────────────────────────────

tab_flow, tab_rules, tab_util, tab_data = st.tabs([
    "🔄 배정 플로우챠트",
    "📋 규칙별 설명 + 배정 결과",
    "🌡️ 작업장 조업도",
    "🗂️ 상세 데이터",
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 : 배정 플로우챠트
# ══════════════════════════════════════════════════════════════════════════════

with tab_flow:
    st.subheader("📌 배정 규칙 전체 플로우챠트")
    st.info(
        "🟡 노란 박스: 리포팅 단계 (배정 없음)  |  "
        "🔵 파란 박스: 글로벌 계산 단계  |  "
        "🟢 초록 박스: 목표 산출 및 저장"
    )

    mermaid_html = """<!DOCTYPE html>
<html>
<head>
<script type="module">
  import mermaid from 'https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.esm.min.mjs';
  mermaid.initialize({ startOnLoad: true, theme: 'default', flowchart: { curve: 'basis' } });
</script>
<style>body{margin:0;padding:8px;background:transparent;}.mermaid{max-width:100%;overflow-x:auto;}</style>
</head>
<body>
<div class="mermaid">
flowchart TD
    START([▶ 시작]) --> LOAD[📂 데이터 로드<br/>blk_master + area_capa]
    LOAD --> CALC[⚙️ 글로벌: load_mh 부하 분산<br/>m_stdt~m_fndt 영업일 균등 배분<br/>→ 월별 부하 + 주별 부하 집계]
    CALC --> R0{📊 R0 이관물량 처리가능 리포팅<br/>T부하 + H부하×1.2 vs 전체능력<br/>초과 시 월별 경고 출력}
    R0 --> R1[📌 R1 목표물량 산출<br/>목표조업도 = 전체부하 ÷ 전체능력<br/>배정목표 = 목표조업도 × 작업장능력]
    R1 --> R2[R2 H행 + blk Y9시작<br/>→ AREA15 무조건 배정]
    R2 --> R3[R3 T행 + stg=중·대<br/>→ m_area 배정]
    R3 --> R4[R4 T행 + stg=소 + area_prior_1=동일<br/>→ prt_area 배정]
    R4 --> R5{📊 R5 H-소 처리가능 리포팅<br/>H-소 부하 vs AREA3~5 잔여능력<br/>초과 시 월별 경고 출력}
    R5 --> R6[R6 H행 + stg=소<br/>AREA3 → AREA4 → AREA5 순환 배정<br/>목표물량 도달 시 해당 area 중단]
    R6 --> R7[R7-Part1 H+중+ship_kind=A+blk∈D114·D174·D204<br/>→ AREA2 무조건 배정<br/>R7-Part2 H+중+jig=F<br/>→ AREA2 목표물량까지 배정]
    R7 --> R8[R8 H행 + stg=중 + jig=D·L<br/>단 H·E11·E51 시작 블록 제외<br/>AREA1 → AREA2 → AREA6 순환 배정]
    R8 --> R9[R9 H행 + stg=대·중 전문화 Family<br/>1-1: blk H시작 + 말미P → AREA7<br/>1-2: blk H시작 + 말미S → AREA8<br/>2-1: blk E11·F51시작 + 말미P → AREA9<br/>2-2: blk E11·F51시작 + 말미S → AREA10]
    R9 --> R10[R10 H행 + stg=대 + jig=D·L<br/>동일 그룹 포함하여 → AREA12<br/>목표물량 도달 시 중단]
    R10 --> R11[R11 H행 + stg=대·중 잔여물량<br/>PS페어 그룹: AREA7+8 · AREA9+10 · AREA13+14<br/>단독 그룹: AREA7→14 순차 first-fit<br/>그룹 초과 시 해제 후 개별 배정]
    R11 --> R12[R12 H행 + blk 4번째 자리 9 또는 H<br/>→ prt_area로 override 배정]
    R12 --> R13[R13 T행 + stg=소 미배정 잔여<br/>area_seq 순으로 분산 배정<br/>그룹: pjt+blk+type / 정렬: assign_prior→m_stdt<br/>area_prior_1→5 단계적 확장 / 목표 도달 시 다음 area]
    R13 --> SAVE[💾 결과 저장 완료]
    style R0 fill:#fff3cd,stroke:#ffc107
    style R5 fill:#fff3cd,stroke:#ffc107
    style CALC fill:#d1ecf1,stroke:#0c7cd5
    style R1 fill:#d4edda,stroke:#28a745
    style SAVE fill:#d4edda,stroke:#28a745
</div>
</body>
</html>"""
    components.html(mermaid_html, height=900, scrolling=True)

    st.divider()
    st.subheader("규칙별 한줄 요약")
    cols = st.columns(2)
    for i, (rule, meta) in enumerate(RULE_META.items()):
        bg = BADGE_BG[meta["badge"]]
        fg = BADGE_COLOR[meta["badge"]]
        with cols[i % 2]:
            st.markdown(
                f"""<div style="background:{bg};border-left:4px solid {fg};
                    padding:8px 12px;border-radius:4px;margin-bottom:8px;font-size:13px">
                    <b style="color:{fg}">{rule}</b>&nbsp; {meta['label']}<br/>
                    <span style="color:#555">{meta['desc']}</span>
                </div>""",
                unsafe_allow_html=True,
            )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 : 규칙별 설명 + 배정 결과
# ══════════════════════════════════════════════════════════════════════════════

with tab_rules:
    steps = stats["steps"]
    steps_map = {s["rule"]: s for s in steps}

    st.subheader("규칙별 배정 블록 수 전체 현황")
    rule_df = pd.DataFrame([{"rule": s["rule"], "label": s["label"], "count": s["count"]} for s in steps])
    fig_all = px.bar(
        rule_df, x="rule", y="count", text="count",
        hover_data=["label"], color="count", color_continuous_scale="Blues",
        labels={"rule": "규칙", "count": "블록 수"},
    )
    fig_all.update_traces(textposition="outside")
    fig_all.update_layout(coloraxis_showscale=False, height=350)
    st.plotly_chart(fig_all, use_container_width=True)

    st.subheader("규칙별 작업장 분포 히트맵")
    rule_area_rows = []
    for s in steps:
        for area, cnt in s["areas"].items():
            rule_area_rows.append({"rule": s["rule"], "area": area, "count": cnt})
    rule_area_df = pd.DataFrame(rule_area_rows)
    if not rule_area_df.empty:
        pivot = rule_area_df.pivot_table(index="area", columns="rule", values="count", fill_value=0)
        pivot = pivot[sorted(pivot.columns, key=lambda x: int(x[1:]) if x[1:].isdigit() else 0)]
        pivot = pivot.reindex(sorted(pivot.index, key=area_key))
        fig_heat = px.imshow(
            pivot, text_auto=True, color_continuous_scale="YlOrRd",
            labels={"x": "규칙", "y": "작업장", "color": "블록 수"}, aspect="auto",
        )
        fig_heat.update_layout(height=480)
        st.plotly_chart(fig_heat, use_container_width=True)

    st.divider()
    st.subheader("규칙별 상세 설명 + 배정 결과")

    cumul = 0
    for rule, meta in RULE_META.items():
        step  = steps_map.get(rule)
        count = step["count"] if step else 0
        cumul += count
        bg = BADGE_BG[meta["badge"]]
        fg = BADGE_COLOR[meta["badge"]]

        with st.expander(f"{rule}  ·  {meta['label']}  ({count:,}건)", expanded=False):
            st.markdown(
                f"""<div style="background:{bg};border-left:4px solid {fg};
                    padding:10px 14px;border-radius:4px;margin-bottom:12px">
                    <b style="color:{fg};font-size:14px">{rule} — {meta['label']}</b><br/>
                    <span style="font-size:13px;color:#333">{meta['detail']}</span>
                </div>""",
                unsafe_allow_html=True,
            )
            col_l, col_r = st.columns([1, 1])
            with col_l:
                st.markdown("**배정 조건 요약**")
                st.caption(meta["desc"])
                st.markdown(f"- 배정 건수: **{count:,}건**")
                st.markdown(f"- 누적 배정: **{cumul:,}건**")
            with col_r:
                if step and step["areas"]:
                    area_rows = sorted(step["areas"].items(), key=lambda x: area_key(x[0]))
                    area_df = pd.DataFrame(area_rows, columns=["작업장", "블록수"])
                    fig_r = px.bar(
                        area_df, x="작업장", y="블록수", text="블록수",
                        color="블록수", color_continuous_scale="Blues", height=260,
                    )
                    fig_r.update_traces(textposition="outside")
                    fig_r.update_layout(coloraxis_showscale=False, margin=dict(t=20, b=20, l=0, r=0))
                    st.plotly_chart(fig_r, use_container_width=True)
                else:
                    st.info("배정 없음 (리포팅 전용 단계)")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 : 작업장 조업도
# ══════════════════════════════════════════════════════════════════════════════

with tab_util:
    st.subheader("작업장별 월별 조업도 (%)")

    util_data = stats["utilization"]
    util_df = (
        pd.DataFrame(util_data).T
        .reset_index().rename(columns={"index": "area"})
        .melt(id_vars="area", var_name="month", value_name="util")
    )
    util_df["util"] = pd.to_numeric(util_df["util"], errors="coerce").round(1)
    util_df = util_df.sort_values(["area", "month"], key=lambda s: s.map(area_key) if s.name == "area" else s)

    pivot_util = util_df.pivot(index="area", columns="month", values="util")
    pivot_util = pivot_util[sorted(pivot_util.columns)]
    pivot_util = pivot_util.reindex(sorted(pivot_util.index, key=area_key))

    fig_util = px.imshow(
        pivot_util, text_auto=".1f",
        color_continuous_scale=[(0.00,"#aec6e8"),(0.50,"#4caf50"),(0.75,"#ffc107"),(1.00,"#f44336")],
        zmin=0, zmax=200, labels={"x": "월", "y": "작업장", "color": "조업도(%)"}, aspect="auto",
    )
    fig_util.update_layout(height=500)
    st.plotly_chart(fig_util, use_container_width=True)

    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("작업장 선택 — 월별 추이")
        sel_areas_chart = st.multiselect(
            "작업장",
            options=sorted(util_data.keys(), key=area_key),
            default=sorted(util_data.keys(), key=area_key)[:4],
            key="util_area_select",
        )
        if sel_areas_chart:
            fig_line = px.line(
                util_df[util_df["area"].isin(sel_areas_chart)],
                x="month", y="util", color="area", markers=True,
                labels={"month": "월", "util": "조업도(%)", "area": "작업장"},
            )
            fig_line.add_hline(y=100, line_dash="dash", line_color="red", annotation_text="100%")
            st.plotly_chart(fig_line, use_container_width=True)

    with col_b:
        st.subheader("월 선택 — 작업장별 비교")
        months = sorted(util_df["month"].unique())
        sel_month = st.selectbox("월", months, index=0)
        month_df = util_df[util_df["month"] == sel_month].sort_values("area", key=lambda s: s.map(area_key))
        fig_bar = px.bar(
            month_df, x="area", y="util", text="util",
            color="util", color_continuous_scale="RdYlGn_r",
            labels={"area": "작업장", "util": "조업도(%)"},
        )
        fig_bar.add_hline(y=100, line_dash="dash", line_color="red")
        fig_bar.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
        fig_bar.update_layout(coloraxis_showscale=False, height=350)
        st.plotly_chart(fig_bar, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 : 상세 데이터
# ══════════════════════════════════════════════════════════════════════════════

with tab_data:
    st.subheader(f"배정 결과 데이터 ({len(df_filtered):,}건)")

    col_f1, col_f2 = st.columns([3, 1])
    with col_f1:
        rule_opts = ["전체"] + sorted(
            df_all["rule_assigned"].dropna().unique().tolist(),
            key=lambda x: int(x[1:]) if x[1:].isdigit() else 0,
        )
        sel_rule = st.selectbox("규칙 필터", rule_opts)
    with col_f2:
        show_unassigned = st.checkbox("미배정만 보기", value=False)

    view_df = df_filtered.copy()
    if show_unassigned:
        view_df = view_df[view_df["assigned_area"].isna()]
    elif sel_rule != "전체":
        view_df = view_df[view_df["rule_assigned"] == sel_rule]

    for col in ["m_stdt", "m_fndt"]:
        view_df[col] = view_df[col].dt.strftime("%Y-%m-%d")

    display_cols = [
        "pjt", "ship_kind", "h_t", "blk", "stg", "type", "jig",
        "wt", "m_stdt", "m_fndt", "m_area", "prt_area",
        "load_mh", "assigned_area", "rule_assigned",
    ]
    st.dataframe(view_df[display_cols].reset_index(drop=True), use_container_width=True, height=520)

    csv = view_df[display_cols].to_csv(index=False, encoding="utf-8-sig")
    st.download_button("CSV 다운로드", data=csv, file_name="blk_assign_filtered.csv", mime="text/csv")
